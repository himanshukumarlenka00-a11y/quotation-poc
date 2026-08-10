"""Guards the money maths on a customer-facing quotation.

History worth knowing before you change this file. The export wrote Excel
formulas; fd0d76b (2026-08-04) replaced them with values because manual-calc
Excel showed the cells BLANK; 8f33e05 (2026-08-10) put formulas back, on
request, so the sheet recalculates when someone edits a price — and that is
only safe because the workbook now carries fullCalcOnLoad, which forces Excel
to recalculate on open. If that flag ever goes, the 2026-08-04 bug returns.

So the money is in formulas again, and a test that only read cell values
would assert nothing. These tests EVALUATE the formulas — same arithmetic
Excel will do — so a wrong reference or a wrong range still fails here rather
than on a customer's desk.

Plain asserts, no framework:  python test_pricing.py
"""
import re

import openpyxl

from app.export import build_company_quotation

# Column letters in the company template (see build_company_quotation).
QTY, PRICE, AMOUNT, GST_PCT, GST_VAL = "C", "J", "K", "L", "M"
FIRST_ITEM_ROW = 21


def ev(ws, ref, _depth=0):
    """Evaluate one cell, following formula references. Handles exactly the
    four shapes the exporter emits: =C21*J21, =K21*L21, =SUM(K21:K38), =K39+K40.
    Anything else is a formula this test does not know about — fail loudly
    rather than silently return None and let an assert pass by accident."""
    assert _depth < 12, f"formula loop at {ref}"
    v = ws[ref].value
    if not (isinstance(v, str) and v.startswith("=")):
        return v or 0
    body = v[1:]
    m = re.fullmatch(r"SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)", body)
    if m:
        col, r1, r2 = m.group(1), int(m.group(2)), int(m.group(4))
        return round(sum(ev(ws, f"{col}{r}", _depth + 1) for r in range(r1, r2 + 1)), 2)
    m = re.fullmatch(r"([A-Z]+\d+)([*+])([A-Z]+\d+)", body)
    if m:
        a, op, b = ev(ws, m.group(1), _depth + 1), m.group(2), ev(ws, m.group(3), _depth + 1)
        return round(a * b if op == "*" else a + b, 2)
    raise AssertionError(f"unrecognised formula in {ref}: {v}")


def _build(items, **quote):
    q = {"ref_no": "TEST-PRICING", "client_name": "Test"}
    q.update(quote)
    return openpyxl.load_workbook(build_company_quotation(q, items))["QUOTE"]


def item(**kw):
    base = {"product": "P", "qty": 1, "price_per_pc": 100.0, "gst_pct": 18.0,
            "model_no": "M", "brand": "B", "hsn_code": "1234", "specification": "s",
            "description": "d"}
    base.update(kw)
    return base


def test_line_amount_and_gst():
    ws = _build([item(qty=4, price_per_pc=318.0, gst_pct=18.0)])
    r = FIRST_ITEM_ROW
    assert ws[f"{QTY}{r}"].value == 4
    assert ws[f"{PRICE}{r}"].value == 318.0
    assert ws[f"{GST_PCT}{r}"].value == 0.18, "GST% must be written as a decimal fraction"
    assert ev(ws, f"{AMOUNT}{r}") == 1272.0, f"qty*price wrong: {ev(ws, AMOUNT + str(r))}"
    assert ev(ws, f"{GST_VAL}{r}") == 228.96, f"amount*gst wrong: {ev(ws, GST_VAL + str(r))}"


def test_zero_gst_stays_zero():
    """A genuine 0% product must not silently become 18% -- this shipped once."""
    ws = _build([item(qty=2, price_per_pc=50.0, gst_pct=0.0)])
    r = FIRST_ITEM_ROW
    assert ev(ws, f"{AMOUNT}{r}") == 100.0
    assert ws[f"{GST_PCT}{r}"].value == 0.0, f"0% GST flipped to {ws[GST_PCT + str(r)].value}"
    assert ev(ws, f"{GST_VAL}{r}") == 0.0, "0% GST produced tax"


def test_the_editable_cells_are_real_numbers():
    """QTY / PRICE/PC / GST% are what a human types over. If any of them
    became a formula, editing the sheet would fight back."""
    ws = _build([item(qty=3, price_per_pc=99.99)])
    for col in (QTY, PRICE, GST_PCT):
        v = ws[f"{col}{FIRST_ITEM_ROW}"].value
        assert not (isinstance(v, str) and v.startswith("=")), f"{col} must stay typed-in: {v}"


def test_recalculates_on_open():
    """The whole reason formulas are allowed back (see the module docstring).
    Without this flag, manual-calc Excel shows every derived cell blank."""
    wb = openpyxl.load_workbook(build_company_quotation(
        {"ref_no": "TEST-CALC", "client_name": "T"}, [item()]))
    assert wb.calculation.fullCalcOnLoad is True, \
        "fullCalcOnLoad is off — formula cells will render blank in manual-calc Excel"


def test_totals_match_the_lines():
    items = [item(qty=4, price_per_pc=318.0, gst_pct=18.0),
             item(qty=2, price_per_pc=50.0, gst_pct=0.0),
             item(qty=1, price_per_pc=1010.0, gst_pct=12.0)]
    ws = _build(items)

    total_row = FIRST_ITEM_ROW + len(items) + 1   # +1 for the freight row
    gst_row, grand_row = total_row + 1, total_row + 2

    expected_amount = 4 * 318.0 + 2 * 50.0 + 1 * 1010.0            # 2382.0
    expected_gst = round(1272.0 * .18, 2) + 0.0 + round(1010.0 * .12, 2)

    assert ev(ws, f"{AMOUNT}{total_row}") == expected_amount, ev(ws, AMOUNT + str(total_row))
    assert ev(ws, f"{GST_VAL}{total_row}") == expected_gst, ev(ws, GST_VAL + str(total_row))
    assert ev(ws, f"{AMOUNT}{gst_row}") == expected_gst
    assert ev(ws, f"{AMOUNT}{grand_row}") == round(expected_amount + expected_gst, 2)


def test_totals_cover_every_line():
    """A SUM range that stops short is the quiet way to under-bill. Check the
    range spans the first item through the freight row, not a fixed guess."""
    items = [item() for _ in range(7)]
    ws = _build(items)
    freight_row = FIRST_ITEM_ROW + len(items)
    total_row = freight_row + 1
    f = ws[f"{AMOUNT}{total_row}"].value
    assert f == f"=SUM({AMOUNT}{FIRST_ITEM_ROW}:{AMOUNT}{freight_row})", f


def test_freight_is_charged_and_taxed():
    items = [item(qty=1, price_per_pc=1000.0, gst_pct=18.0)]
    ws = _build(items, freight_charge=500.0)

    freight_row = FIRST_ITEM_ROW + len(items)
    total_row = freight_row + 1
    grand_row = total_row + 2

    assert ev(ws, f"{AMOUNT}{freight_row}") == 500.0, "freight missing from the sheet"
    freight_gst = ev(ws, f"{GST_VAL}{freight_row}")
    assert ev(ws, f"{AMOUNT}{total_row}") == 1500.0, "freight not added to the total"
    assert ev(ws, f"{AMOUNT}{grand_row}") == round(1500.0 + 180.0 + freight_gst, 2)


def test_rounding_holds_over_many_lines():
    """0.01 drift compounding across a large BOQ would be invisible per line."""
    items = [item(qty=3, price_per_pc=33.33, gst_pct=18.0) for _ in range(100)]
    ws = _build(items)
    total_row = FIRST_ITEM_ROW + len(items) + 1

    line_amount = round(3 * 33.33, 2)                     # 99.99
    assert ev(ws, f"{AMOUNT}{total_row}") == round(line_amount * 100, 2)


if __name__ == "__main__":
    for name, fn in sorted(list(globals().items())):
        if name.startswith("test_") and callable(fn):
            fn()
            print(f"  ok  {name}")
    print("pricing maths OK")
