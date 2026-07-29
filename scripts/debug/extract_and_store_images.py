"""
Extract product images from uploaded .xls catalog files and store them
as base64 in the DB, mapped to each product by row anchor.

Requires LibreOffice (for .xls -> .xlsx conversion).
Run locally once; the DB (with embedded images) is then deployed to Railway.
"""
import os, sqlite3, base64, subprocess, sys, re

BASE = os.path.dirname(os.path.abspath(__file__))
UPLOADS = os.path.join(BASE, "uploads")
CONVERTED = os.path.join(BASE, "data", "converted")
DB = os.path.join(BASE, "data", "quotations.db")
SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"

os.makedirs(CONVERTED, exist_ok=True)

import openpyxl
from PIL import Image
import io


def ensure_image_column():
    conn = sqlite3.connect(DB)
    try:
        conn.execute("ALTER TABLE boq_items ADD COLUMN image_data TEXT DEFAULT ''")
        conn.commit()
        print("Added image_data column")
    except Exception:
        print("image_data column already exists")
    conn.close()


def convert_to_xlsx(xls_path):
    """Convert .xls to .xlsx using LibreOffice."""
    name = os.path.splitext(os.path.basename(xls_path))[0] + ".xlsx"
    out = os.path.join(CONVERTED, name)
    if os.path.exists(out):
        os.remove(out)
    subprocess.run(
        [SOFFICE, "--headless", "--convert-to", "xlsx", "--outdir", CONVERTED, xls_path],
        check=False, capture_output=True, timeout=120
    )
    return out if os.path.exists(out) else None


def find_col(header_row, *names):
    """Find column index (1-based) by header name substring."""
    for idx, val in enumerate(header_row, start=1):
        if val and any(n in str(val).upper() for n in names):
            return idx
    return None


def downscale_image(blob, max_px=200):
    """Resize image to keep base64 small."""
    try:
        img = Image.open(io.BytesIO(blob))
        img = img.convert("RGB")
        img.thumbnail((max_px, max_px))
        out = io.BytesIO()
        img.save(out, format="JPEG", quality=70)
        return out.getvalue()
    except Exception:
        return blob


def process_file(xls_name):
    xls_path = os.path.join(UPLOADS, xls_name)
    print(f"\n=== {xls_name} ===")
    xlsx = convert_to_xlsx(xls_path)
    if not xlsx:
        print("  conversion failed")
        return {}

    wb = openpyxl.load_workbook(xlsx)
    # product_row_map: {(sheet, product_name_upper): base64_data}
    mapping = {}

    for sh in wb.sheetnames:
        ws = wb[sh]
        imgs = getattr(ws, "_images", [])
        if not imgs:
            continue

        # find header row + product column
        product_col = None
        header_row_idx = None
        for r in range(1, min(ws.max_row, 30) + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 20)]
            pc = find_col(row_vals, "PRODUCT NAME", "PRODUCT", "ITEM NAME", "ITEM")
            if pc:
                product_col = pc
                header_row_idx = r
                break
        if not product_col:
            continue

        count = 0
        for im in imgs:
            try:
                row = im.anchor._from.row + 1  # 1-based
            except Exception:
                continue
            if header_row_idx and row <= header_row_idx:
                continue  # skip logo/header images

            product = ws.cell(row=row, column=product_col).value
            if not product or not str(product).strip():
                continue
            pname = str(product).strip().upper()

            # extract image bytes
            try:
                blob = im._data() if callable(im._data) else im.ref
            except Exception:
                continue
            if not blob or len(blob) < 800:
                continue

            small = downscale_image(blob)
            b64 = "data:image/jpeg;base64," + base64.b64encode(small).decode()
            mapping[(sh, pname)] = b64
            count += 1
        if count:
            print(f"  Sheet '{sh}': mapped {count} images")
    return mapping


def main():
    ensure_image_column()
    conn = sqlite3.connect(DB)

    xls_files = [f for f in os.listdir(UPLOADS) if f.endswith(".xls")]
    # only process files that are in the catalog
    catalog_files = set(r[0] for r in conn.execute("SELECT DISTINCT file_name FROM boq_items").fetchall())

    total_updated = 0
    for xls in xls_files:
        if xls not in catalog_files:
            print(f"skip (not in catalog): {xls}")
            continue
        mapping = process_file(xls)
        for (sheet, pname), b64 in mapping.items():
            # match by sheet + product name (case-insensitive)
            cur = conn.execute(
                "UPDATE boq_items SET image_data=? WHERE file_name=? AND UPPER(product)=? AND (sheet_name=? OR sheet_name='' OR sheet_name IS NULL)",
                (b64, xls, pname, sheet)
            )
            if cur.rowcount == 0:
                # fallback: match by name only within file
                cur = conn.execute(
                    "UPDATE boq_items SET image_data=? WHERE file_name=? AND UPPER(product)=?",
                    (b64, xls, pname)
                )
            total_updated += cur.rowcount
    conn.commit()

    with_img = conn.execute("SELECT COUNT(*) FROM boq_items WHERE image_data != ''").fetchone()[0]
    total = conn.execute("SELECT COUNT(*) FROM boq_items").fetchone()[0]
    conn.close()
    print(f"\n=== DONE ===")
    print(f"Products updated with images: {total_updated}")
    print(f"Products with image_data now: {with_img}/{total}")


if __name__ == "__main__":
    main()
