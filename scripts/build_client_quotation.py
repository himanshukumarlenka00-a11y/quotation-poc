"""Fill a client's own requirement workbook (BOQ/REQ) with our pricing.

This produces a CLIENT quotation — distinct from the company quotation that
app/export.py builds. Instead of re-laying the items onto our letterhead, it
keeps the client's workbook exactly as they sent it (their sheets, their row
order, their wording) and appends our offer columns to the right of theirs,
so each row reads "what you asked for | what we're offering".

Matching goes through the app's own resolver (_resolve_master_matches), so a
row here resolves to exactly the same master-table product it would in the
UI — no second, divergent matching implementation to keep in sync.

Usage:
    python scripts/build_client_quotation.py "<req.xlsx>" "<output.xlsx>"
"""
import os
import sys
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.db import get_db                                    # noqa: E402
from app.images import _image_file_path                      # noqa: E402
from app.routers.quotations import _resolve_master_matches   # noqa: E402

# Columns we append to the right of whatever the client's sheet already has.
OUR_COLS = ["DESCRIPTION", "MODEL NO", "BRAND", "IMAGE", "SPECIFICATION",
            "HSN CODE", "PRICE/PC", "AMOUNT", "MATCHED ON"]

HDR_FILL = PatternFill("solid", fgColor="1A3A6B")
HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
THIN = Side(style="thin", color="9BA7B8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '#,##0.00'


def _load_env_key():
    """The resolver needs GROQ_API_KEY; run_local.ps1 injects it from .env."""
    env = Path(__file__).resolve().parent.parent / ".env"
    if env.exists() and not os.environ.get("GROQ_API_KEY"):
        for line in env.read_text(encoding="utf-8").splitlines():
            if line.startswith("GROQ_API_KEY="):
                os.environ["GROQ_API_KEY"] = line.split("=", 1)[1].strip()


def _find_header(ws):
    """Locate the client's header row and map its labels -> column index.

    Sheet-to-sheet the label spacing varies ('Sl. No.' vs 'Sl.No '), so match
    loosely on the serial-number column rather than an exact string.
    """
    for r in range(1, min(12, ws.max_row + 1)):
        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if any(v.lower().startswith("sl") and "no" in v.lower() for v in vals):
            return r, {v: i + 1 for i, v in enumerate(vals) if v}
    return None, {}


def _pick(cols, *names):
    for n in names:
        for label, idx in cols.items():
            if label.strip().lower() == n.lower():
                return idx
    return None


def collect_rows(wb):
    """Every requirement line across every sheet, in workbook order."""
    out = []
    for ws in wb.worksheets:
        hdr, cols = _find_header(ws)
        if not hdr:
            continue
        c_prod = _pick(cols, "Product")
        if not c_prod:
            continue
        c_spec = _pick(cols, "Specification")
        c_model = _pick(cols, "Model No", "Model")
        c_brand = _pick(cols, "Brand")
        c_qty = _pick(cols, "OPM Requirement", "Requirement", "Qty", "Quantity")
        for r in range(hdr + 1, ws.max_row + 1):
            product = str(ws.cell(r, c_prod).value or "").strip()
            if not product:
                continue
            qty = ws.cell(r, c_qty).value if c_qty else None
            try:
                qty = int(float(qty))
            except (TypeError, ValueError):
                qty = 0
            out.append({
                "ws": ws, "row": r, "hdr": hdr,
                "product": product,
                "spec": str(ws.cell(r, c_spec).value or "").strip() if c_spec else "",
                "model": str(ws.cell(r, c_model).value or "").strip() if c_model else "",
                "brand": str(ws.cell(r, c_brand).value or "").strip() if c_brand else "",
                "qty": qty,
            })
    return out


def match_rows(rows, tiers=("3star",)):
    """Resolve every row against the Master Table via the app's resolver.

    The resolver returns only the rows it matched (the rest land in
    not_found), preserving input order — so a single forward walk re-aligns
    its output with our original rows without needing a shared key, which
    matters because product labels repeat across sheets.
    """
    from groq import Groq
    _load_env_key()
    client = Groq(api_key=os.environ.get("GROQ_API_KEY", ""))

    extracted = [{
        "product": r["product"],
        "qty": r["qty"],
        # model + spec give search_catalog the signal to tell same-named
        # rows apart; without it every size of "Plate, Round Rim" collapses
        # onto whichever single master row scores highest for that label.
        "search_term": " ".join(p for p in (r["product"], r["model"], r["spec"]) if p),
    } for r in rows]

    conn = get_db()
    try:
        matched, not_found = _resolve_master_matches(
            conn, extracted, [], list(tiers), client, prompt="")
    finally:
        conn.close()

    j = 0
    for row, ex in zip(rows, extracted):
        if j < len(matched) and matched[j].get("_requested") == ex["product"]:
            row["match"] = matched[j]
            j += 1
        else:
            row["match"] = None
    return rows, not_found


def write_offer(rows):
    """Append our offer columns to each sheet, aligned to the client's rows."""
    seen = {}
    filled = 0
    for row in rows:
        ws, r, hdr = row["ws"], row["row"], row["hdr"]
        if ws.title not in seen:
            start = ws.max_column + 2          # one blank spacer column
            seen[ws.title] = start
            for i, name in enumerate(OUR_COLS):
                cell = ws.cell(hdr, start + i, name)
                cell.fill, cell.font = HDR_FILL, HDR_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = BORDER
                ws.column_dimensions[get_column_letter(start + i)].width = \
                    28 if name in ("DESCRIPTION", "SPECIFICATION") else 14
        start = seen[ws.title]
        m = row["match"]
        if not m:
            c = ws.cell(r, start, "NOT IN MASTER TABLE")
            c.font = Font(italic=True, color="C0392B", size=9)
            c.border = BORDER
            continue

        filled += 1
        price = float(m.get("price_per_pc") or 0)
        amount = price * (row["qty"] or 0)
        basis = "Model No" if row["model"] else "Product + Spec"
        vals = [m.get("product", ""), m.get("model_no", ""), m.get("brand", ""), "",
                m.get("specification", ""), m.get("hsn_code", ""), price, amount, basis]
        for i, v in enumerate(vals):
            cell = ws.cell(r, start + i, v)
            cell.border = BORDER
            cell.font = Font(size=9)
            if i in (6, 7):
                cell.number_format = MONEY
            if i == 8:
                cell.font = Font(size=8, italic=True,
                                 color="1E9E56" if row["model"] else "B8860B")

        img_hash = (m.get("image_path") or "").strip()
        if img_hash:
            path = _image_file_path(img_hash)
            if path and os.path.exists(path):
                try:
                    pic = XLImage(path)
                    pic.width, pic.height = 54, 44
                    ws.add_image(pic, f"{get_column_letter(start + 3)}{r}")
                    ws.row_dimensions[r].height = 36
                except Exception:
                    pass       # a corrupt image must never abort the quote
    return filled


def main(src, dest):
    wb = openpyxl.load_workbook(src)
    rows = collect_rows(wb)
    print(f"requirement rows : {len(rows)}")
    rows, not_found = match_rows(rows)
    filled = write_offer(rows)
    Path(dest).parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)

    total = sum(float(r["match"].get("price_per_pc") or 0) * (r["qty"] or 0)
                for r in rows if r["match"])
    print(f"priced           : {filled}")
    print(f"not in master    : {len(rows) - filled}")
    print(f"quotation value  : Rs {total:,.2f}")
    print(f"saved            : {dest}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
