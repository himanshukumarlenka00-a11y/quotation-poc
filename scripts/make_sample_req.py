"""Generate a sample client requirement workbook from a master catalog.

Used to exercise the client-quotation flow (scripts/build_client_quotation.py)
for catalogs where we don't have a real client REQ file on hand. The layout
mirrors the OPM requirement sheets: the client's own columns only, no pricing.

Model numbers are deliberately left blank on a share of the rows. Real
requirement sheets are like this — in the OPM file only 84 of 711 rows carried
one — and those rows are the ones that force matching to work from the product
and specification text alone, which is what we want to see exercised.

Usage:
    python scripts/make_sample_req.py "<catalog file_name>" <count> "<out.xlsx>"
"""
import os
import sys
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.db import get_db            # noqa: E402
from app.images import _image_file_path   # noqa: E402

HDR = ["Sl. No.", "Product", "Specification", "Brand", "Model No",
       "Image", "Unit", "OPM Requirement", "Uses", "Remarks"]

TITLE_FILL = PatternFill("solid", fgColor="D9E2F3")
HDR_FILL = PatternFill("solid", fgColor="1A3A6B")
THIN = Side(style="thin", color="9BA7B8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Plausible order sizes, cycled so quantities look like a real requirement
# rather than every line asking for the same number.
QTYS = [12, 24, 50, 100, 200, 6, 36, 500, 60, 150, 20, 300, 48, 80, 1000]


def build(catalog, count, dest):
    conn = get_db()
    # Prefer rows that have a photo, so the sample sheet looks like a real
    # client requirement (theirs carry a picture per line) rather than a
    # table of blank Image cells.
    rows = conn.execute("""
        SELECT product, specification, brand, original_model, unit,
               product_group, image_path
          FROM master_products
         WHERE file_name = ? AND product IS NOT NULL AND product <> ''
           AND price_3star > 0
         ORDER BY CASE WHEN image_path IS NOT NULL AND image_path <> ''
                       THEN 0 ELSE 1 END, product
         LIMIT ?
    """, (catalog, count)).fetchall()
    conn.close()
    if not rows:
        raise SystemExit(f"No priced products found for catalog: {catalog}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requirement"

    ws.cell(2, 3, f"REQUIREMENT LIST — {catalog.rsplit('.', 1)[0]}")
    ws.cell(2, 3).font = Font(bold=True, size=12, color="1A3A6B")
    ws.cell(2, 3).fill = TITLE_FILL

    for i, name in enumerate(HDR):
        c = ws.cell(4, 3 + i, name)
        c.fill, c.font = HDR_FILL, Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for n, r in enumerate(rows):
        excel_row = 5 + n
        # Every 5th row keeps its model number; the rest must match on text.
        model = r["original_model"] if n % 5 == 0 else ""
        vals = [n + 1, r["product"], r["specification"] or "", r["brand"] or "",
                model, "", r["unit"] or "Nos", QTYS[n % len(QTYS)],
                r["product_group"] or "", ""]
        for i, v in enumerate(vals):
            c = ws.cell(excel_row, 3 + i, v)
            c.border = BORDER
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="center", wrap_text=i in (1, 2))

        # Client's own Image column (6th) — a real requirement sheet shows the
        # product picture here, which is what makes the row identifiable.
        img = (r["image_path"] or "").strip()
        if img:
            path = _image_file_path(img)
            if path and os.path.exists(path):
                try:
                    pic = XLImage(path)
                    pic.width, pic.height = 56, 44
                    ws.add_image(pic, f"{openpyxl.utils.get_column_letter(3 + 5)}{excel_row}")
                except Exception:
                    pass   # a bad image must never abort sheet generation
        ws.row_dimensions[excel_row].height = 38

    widths = [8, 38, 34, 14, 20, 10, 8, 14, 20, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(3 + i)].width = w

    Path(dest).parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    with_model = sum(1 for n in range(len(rows)) if n % 5 == 0)
    print(f"{len(rows):>4} rows  ({with_model} with model no, "
          f"{len(rows) - with_model} text-only)  ->  {dest}")


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)
    build(sys.argv[1], int(sys.argv[2]), sys.argv[3])
