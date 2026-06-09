import os, json, sqlite3, re, shutil, io, base64
from datetime import datetime
from pathlib import Path
from fastapi import FastAPI, UploadFile, File, HTTPException, Response
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
import pandas as pd
from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import warnings
warnings.filterwarnings("ignore")

BASE = Path(__file__).parent
DATA_DIR = BASE / "data"
UPLOADS_DIR = BASE / "uploads"
EXPORTS_DIR = BASE / "exports"
IMAGES_DIR = BASE / "data" / "images"
for d in [DATA_DIR, UPLOADS_DIR, EXPORTS_DIR, IMAGES_DIR]:
    d.mkdir(parents=True, exist_ok=True)

DB_PATH = DATA_DIR / "quotations.db"
app = FastAPI(title="QuoteGen AI")
app.mount("/static", StaticFiles(directory=BASE / "static"), name="static")


# ── DB ────────────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS boq_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT,
            product TEXT,
            description TEXT,
            model_no TEXT,
            brand TEXT,
            specification TEXT,
            hsn_code TEXT,
            price REAL,
            price_currency TEXT DEFAULT 'INR',
            gst_pct REAL,
            image_path TEXT,
            uploaded_at TEXT
        );
        CREATE TABLE IF NOT EXISTS templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT,
            file_path TEXT,
            structure_json TEXT,
            uploaded_at TEXT
        );
        CREATE TABLE IF NOT EXISTS quotations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ref_no TEXT,
            client_name TEXT,
            items_json TEXT,
            status TEXT DEFAULT 'draft',
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            quotation_id INTEGER,
            rating TEXT,
            missing_items TEXT,
            created_at TEXT
        );
    """)
    conn.commit()
    conn.close()


init_db()

# Add new columns to existing DBs that predate schema changes
def migrate_db():
    conn = get_db()
    for col, definition in [
        ("sheet_name", "TEXT DEFAULT ''"),
        ("price_currency", "TEXT DEFAULT 'INR'"),
    ]:
        try:
            conn.execute(f"ALTER TABLE boq_items ADD COLUMN {col} {definition}")
            conn.commit()
        except Exception:
            pass
    conn.close()

migrate_db()


# ── Dynamic Template Analyzer ─────────────────────────────────────────────────

def analyze_template(wb: openpyxl.Workbook) -> dict:
    """
    Dynamically detect the structure of ANY quotation Excel template.
    Returns structure dict describing where each section lives.
    """
    ws = wb.active
    structure = {
        "header_rows": [],       # rows above the items table (company info)
        "bill_to_row": None,     # row index of "BILL & SHIP TO"
        "client_row": None,      # row where client name goes
        "date_row": None,        # row where date is
        "date_col": None,        # column where date is
        "ref_row": None,         # row where REF NO is
        "salutation_rows": [],   # Dear Sir + intro text rows
        "table_header_row": None, # row with SL.NO / PRODUCT / QTY etc.
        "data_start_row": None,
        "data_end_row": None,    # last actual data row (estimated from template)
        "footer_rows": [],       # rows after data (totals, T&C, bank, signature)
        "col_map": {},           # normalized column name -> 1-based col index
        "max_col": 1,
    }

    nrows = ws.max_row
    ncols = ws.max_column

    def cell_val(r, c):
        v = ws.cell(r, c).value
        return str(v).strip() if v is not None else ""

    def row_text(r):
        return " ".join(cell_val(r, c) for c in range(1, ncols + 1)).strip()

    # Find table header row (contains SL.NO or equivalent)
    for r in range(1, min(nrows + 1, 40)):
        txt = row_text(r).upper()
        if any(k in txt for k in ["SL.NO", "S.NO", "SL NO", "SLNO", "SR.NO"]):
            structure["table_header_row"] = r
            # Map columns
            for c in range(1, ncols + 1):
                v = cell_val(r, c).upper()
                if v:
                    structure["col_map"][v] = c
            structure["data_start_row"] = r + 1
            structure["max_col"] = ncols
            break

    if not structure["table_header_row"]:
        return structure

    thr = structure["table_header_row"]

    # Scan above header row for sections
    for r in range(1, thr):
        txt = row_text(r).upper()
        if "BILL" in txt and "SHIP" in txt:
            structure["bill_to_row"] = r
            # Date is usually on the same row but far right
            for c in range(ncols, 0, -1):
                v = cell_val(r, c)
                if "DATE" in v.upper():
                    structure["date_row"] = r
                    structure["date_col"] = c
                    break
            structure["client_row"] = r + 1
        if "REF" in txt and ("NO" in txt or ":" in txt):
            structure["ref_row"] = r
        structure["header_rows"].append(r)

    # Find data end row and footer
    in_data = False
    for r in range(thr + 1, nrows + 1):
        txt = row_text(r).upper()
        first = cell_val(r, 1).upper()
        # Stop at totals/footer markers
        if any(marker in first for marker in ["TOTAL", "GST VALUE", "GRAND TOTAL", "TERMS", "BANK", "CHEQUE"]):
            structure["data_end_row"] = r - 1
            structure["footer_rows"] = list(range(r, nrows + 1))
            break
        in_data = True

    if in_data and not structure["data_end_row"]:
        structure["data_end_row"] = nrows

    return structure


# ── Image Extraction ──────────────────────────────────────────────────────────

def extract_images_from_wb(wb: openpyxl.Workbook, structure: dict) -> dict:
    """
    Extract embedded images and map them to data row indices.
    Returns {relative_row_in_data: image_bytes}
    """
    ws = wb.active
    img_map = {}
    thr = structure.get("table_header_row", 0)
    for img in getattr(ws, "_images", []):
        try:
            anchor = img.anchor
            # twoCellAnchor or oneCellAnchor
            if hasattr(anchor, '_from'):
                row_0idx = anchor._from.row   # 0-indexed
                row_1idx = row_0idx + 1        # 1-indexed
            elif hasattr(anchor, 'row'):
                row_1idx = anchor.row
            else:
                continue
            if row_1idx > thr:
                rel_row = row_1idx - thr  # 1 = first data row
                img_bytes = img._data() if callable(img._data) else img._data
                img_map[rel_row] = img_bytes
        except Exception:
            pass
    return img_map


def save_product_image(image_bytes: bytes, product: str, model_no: str) -> str:
    """Save image bytes to disk, return relative path."""
    safe = re.sub(r'[^\w\-]', '_', f"{product}_{model_no}")[:60]
    path = IMAGES_DIR / f"{safe}.png"
    with open(str(path), "wb") as f:
        f.write(image_bytes)
    return str(path)


# ── BOQ Parser ────────────────────────────────────────────────────────────────

def parse_boq_excel(filepath: str, filename: str):
    """
    Fully dynamic parser — detects structure from the file itself.
    Returns (items, structure_dict, image_map)
    """
    items = []
    img_map = {}
    structure = {}

    # openpyxl only works with .xlsx — use it for images/structure when possible
    if filepath.endswith(".xlsx"):
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            structure = analyze_template(wb)
            img_map = extract_images_from_wb(wb, structure)
        except Exception:
            structure = {}
            img_map = {}

    # Load all sheets with pandas
    try:
        xl = pd.ExcelFile(filepath)
        sheet_names = xl.sheet_names
    except Exception:
        return items, structure, img_map

    skip_keywords = {"TOTAL", "GRAND TOTAL", "GST VALUE", "ADD GST", "NAN", "", "OPTION",
                     "SUMMARY", "TERMS", "BANK", "SI NO", "SHEET NAME"}

    for sheet_name in sheet_names:
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        except Exception:
            continue

        # Detect header row for this sheet independently
        sheet_thr = None
        sheet_col_map = {}
        for i, row in df.iterrows():
            vals = [str(v).upper().strip() for v in row.values if pd.notna(v)]
            if any(k in vals for k in ["SL.NO", "S.NO", "SLNO", "PRODUCT", "ITEM"]):
                sheet_thr = i
                for ci_, v in enumerate(row.values):
                    if pd.notna(v) and str(v).strip():
                        sheet_col_map[str(v).upper().strip()] = ci_
                break

        if sheet_thr is None:
            continue

        # Use structure from openpyxl for first sheet if available
        if not structure:
            structure["table_header_row"] = sheet_thr + 1
            structure["col_map"] = {k: v+1 for k, v in sheet_col_map.items()}

        def find_col(*names, exclude=None):
            """Find column index by matching name substring. exclude= list of substrings to avoid."""
            for n in names:
                for key, idx in sheet_col_map.items():
                    if n in key:
                        if exclude and any(ex in key for ex in exclude):
                            continue
                        return idx
            return None

        ci = {
            "product":       find_col("PRODUCT NAME", "PRODUCT", "ITEM NAME", "ITEM"),
            "qty":           find_col("TOTAL QTY", "QTY", "QUANTITY"),
            "description":   find_col("DESCRIPTION", "DESC"),
            "model_no":      find_col("MODEL NO", "MODEL"),
            "brand":         find_col("BRAND", "MAKE"),
            "specification": find_col("OUR SPECIFICATION", "SPECIFICATION", "SPEC"),
            "hsn_code":      find_col("HSN"),
            # INR price: explicitly exclude any column containing USD
            "price_inr":     find_col("PRICE/PC INR", "PRICE INR", "PRICE/PC", "PRICE", "RATE", "UNIT PRICE", exclude=["USD", "DOLLAR"]),
            # USD price: must contain USD or DOLLAR
            "price_usd":     find_col("PRICE/PC USD", "UNIT PRICE USD", "PRICE USD", "AMOUNT USD", "TOTAL USD"),
            "gst_pct":       find_col("GST%", "GST"),
        }

        if ci["product"] is None:
            continue

        for df_row_idx, row in df.iloc[sheet_thr + 1:].iterrows():
            vals = row.values

            product = vals[ci["product"]] if ci["product"] is not None else None
            if product is None or (isinstance(product, float) and pd.isna(product)):
                continue
            product_str = str(product).strip()
            if not product_str or product_str.upper() in skip_keywords:
                continue
            # Skip rows that look like headers or section titles (no price)
            price_col = ci.get("price_inr") or ci.get("price_usd") or ci.get("price")
            if price_col is not None:
                price_val = vals[price_col]
                if isinstance(price_val, str) and not price_val.replace('.','').replace(',','').isdigit():
                    continue

            def g(key):
                idx = ci.get(key)
                if idx is None: return ""
                v = vals[idx]
                if isinstance(v, float) and pd.isna(v): return ""
                return str(v).strip()

            def gn(key, default=0.0):
                idx = ci.get(key)
                if idx is None: return default
                try:
                    v = vals[idx]
                    return float(v) if pd.notna(v) else default
                except: return default

            rel_row = df_row_idx - sheet_thr
            img_path = ""
            if rel_row in img_map:
                try:
                    img_path = save_product_image(img_map[rel_row], product_str, g("model_no"))
                except Exception:
                    pass

            items.append({
                "file_name": filename,
                "product": product_str,
                "description": g("description"),
                "model_no": g("model_no"),
                "brand": g("brand"),
                "specification": g("specification"),
                "hsn_code": g("hsn_code"),
                "price": gn("price_inr") or gn("price_usd") or gn("price"),
                "price_currency": "INR" if gn("price_inr") else ("USD" if gn("price_usd") else "INR"),
                "gst_pct": (lambda g: round(g * 100, 2) if g and g <= 1 else (g or 18.0))(gn("gst_pct")),
                "image_path": img_path,
                "uploaded_at": datetime.now().isoformat(),
                "sheet_name": sheet_name,
            })

    return items, structure, img_map


# ── Template-Based XLS Builder ────────────────────────────────────────────────

def copy_cell_style(src, dst):
    """Copy formatting from one cell to another."""
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def build_xls_from_template(template_path: str, structure: dict,
                             quotation: dict, items: list) -> str:
    """
    Build output XLS by cloning the actual template file and filling in data.
    Preserves ALL formatting, merged cells, fonts, colors from original.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    thr = structure.get("table_header_row", 1)
    data_start = thr + 1
    footer_rows = structure.get("footer_rows", [])
    footer_start = min(footer_rows) if footer_rows else ws.max_row + 1

    # ── Update variable header fields ────────────────────────────────────────
    # Client name
    client_row = structure.get("client_row")
    if client_row:
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(client_row, c).value or "")
            if v.strip() and "BILL" not in v.upper() and "SHIP" not in v.upper():
                ws.cell(client_row, c).value = quotation.get("client_name", "")
                break

    # Date
    date_row = structure.get("date_row")
    date_col = structure.get("date_col")
    if date_row and date_col:
        cell = ws.cell(date_row, date_col)
        existing = str(cell.value or "")
        prefix = existing.split(":")[0] + ": " if ":" in existing else "DATE : "
        cell.value = prefix + quotation.get("date", datetime.now().strftime("%d-%m-%Y"))

    # Ref No
    ref_row = structure.get("ref_row")
    if ref_row:
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(ref_row, c).value or "")
            if "REF" in v.upper():
                # Replace just the value part after ":"
                ws.cell(ref_row, c).value = f"REF NO: {quotation.get('ref_no', '')}"
                break

    # ── Remove existing data rows (between header and footer) ────────────────
    rows_to_delete = list(range(data_start, footer_start))
    if rows_to_delete:
        ws.delete_rows(data_start, len(rows_to_delete))

    # Get style from original first data row (before deletion) as template
    # We already deleted, so we'll build style manually using header row as reference
    col_map = structure.get("col_map", {})

    def find_col_idx(*names):
        for n in names:
            for key, idx in col_map.items():
                if n in key:
                    return idx
        return None

    ci = {
        "sl":     find_col_idx("SL.NO", "S.NO", "SL NO", "SLNO"),
        "prod":   find_col_idx("PRODUCT"),
        "qty":    find_col_idx("QTY", "QUANTITY"),
        "desc":   find_col_idx("DESCRIPTION"),
        "model":  find_col_idx("MODEL NO", "MODEL"),
        "brand":  find_col_idx("BRAND"),
        "image":  find_col_idx("IMAGE"),
        "spec":   find_col_idx("SPECIFICATION", "SPEC"),
        "hsn":    find_col_idx("HSN"),
        "price":  find_col_idx("PRICE/PC", "PRICE", "RATE"),
        "amount": find_col_idx("AMOUNT"),
        "gst":    find_col_idx("GST%", "GST"),
        "gstval": find_col_idx("GST VALUE", "GST VAL"),
    }

    max_col = structure.get("max_col", 8)

    # Reference style from header row
    ref_row_style = {c: ws.cell(thr, c) for c in range(1, max_col + 1)}

    # ── Insert new data rows ──────────────────────────────────────────────────
    for idx, item in enumerate(items):
        r = data_start + idx
        ws.insert_rows(r)

        price = float(item.get("price_per_pc") or item.get("price") or 0)
        qty = int(item.get("qty") or 0)
        gst_pct = float(item.get("gst_pct") or 18)
        amount = qty * price
        gst_val = amount * gst_pct / 100

        data = {
            ci["sl"]:     item.get("sl_no", idx + 1),
            ci["prod"]:   item.get("product", ""),
            ci["qty"]:    qty,
            ci["desc"]:   item.get("description", ""),
            ci["model"]:  item.get("model_no", ""),
            ci["brand"]:  item.get("brand", ""),
            ci["image"]:  "",
            ci["spec"]:   item.get("specification", "").replace("\\n", "\n"),
            ci["hsn"]:    item.get("hsn_code", ""),
            ci["price"]:  price,
            ci["amount"]: amount,
            ci["gst"]:    gst_pct / 100,
            ci["gstval"]: gst_val,
        }

        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            copy_cell_style(ref_row_style[c], cell)
            cell.value = data.get(c, "")

        # Number formats for numeric columns
        for col_key, fmt in [("price", "₹#,##0.00"), ("amount", "₹#,##0.00"),
                              ("gst", "0%"), ("gstval", "₹#,##0.00")]:
            col_idx = ci.get(col_key)
            if col_idx:
                ws.cell(r, col_idx).number_format = fmt

        # Embed product image if available
        img_path = item.get("image_path", "")
        if img_path and Path(img_path).exists() and ci.get("image"):
            try:
                xl_img = XLImage(img_path)
                xl_img.width = 80
                xl_img.height = 60
                cell_addr = f"{get_column_letter(ci['image'])}{r}"
                ws.add_image(xl_img, cell_addr)
            except Exception:
                pass

        ws.row_dimensions[r].height = 65

    # ── Update totals in footer ───────────────────────────────────────────────
    total_amount = sum(int(i.get("qty") or 0) * float(i.get("price_per_pc") or i.get("price") or 0) for i in items)
    total_gst = sum(int(i.get("qty") or 0) * float(i.get("price_per_pc") or i.get("price") or 0) * float(i.get("gst_pct") or 18) / 100 for i in items)
    grand_total = total_amount + total_gst

    # Find and update total cells in footer (scan for TOTAL keyword)
    new_footer_start = data_start + len(items)
    for r in range(new_footer_start, ws.max_row + 1):
        for c in range(1, max_col + 1):
            v = str(ws.cell(r, c).value or "").upper().strip()
            if v == "TOTAL":
                # Put total amount in amount column
                if ci.get("amount"):
                    ws.cell(r, ci["amount"]).value = total_amount
                    ws.cell(r, ci["amount"]).number_format = "₹#,##0.00"
                if ci.get("gstval"):
                    ws.cell(r, ci["gstval"]).value = total_gst
                    ws.cell(r, ci["gstval"]).number_format = "₹#,##0.00"
            elif v in ("GST VALUE", "ADD GST@18%", "ADD GST"):
                if ci.get("amount"):
                    ws.cell(r, ci["amount"]).value = total_gst
                    ws.cell(r, ci["amount"]).number_format = "₹#,##0.00"
            elif "GRAND TOTAL" in v:
                if ci.get("amount"):
                    ws.cell(r, ci["amount"]).value = grand_total
                    ws.cell(r, ci["amount"]).number_format = "₹#,##0.00"

    out = EXPORTS_DIR / f"Quote_{quotation.get('ref_no','export').replace('/','-')}.xlsx"
    wb.save(str(out))
    return str(out)


def build_xls_minimal(quotation: dict, items: list) -> str:
    """Build a professional invoice-style Excel quotation."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QUOTATION"

    # ── Colour palette ──
    NAVY   = "1F3864"
    GOLD   = "C8860C"
    LIGHT  = "DCE6F1"
    GREEN  = "E2EFDA"
    WHITE  = "FFFFFF"
    DARK   = "1F3864"

    def cell(r, c, val="", bold=False, size=10, color="000000", bg=None,
             align="left", valign="center", wrap=False, num_fmt=None, italic=False):
        cl = ws.cell(r, c, val)
        cl.font      = Font(bold=bold, size=size, color=color, name="Calibri", italic=italic)
        cl.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
        if bg:
            cl.fill = PatternFill("solid", start_color=bg)
        if num_fmt:
            cl.number_format = num_fmt
        return cl

    def border_row(r, cols, sides="bottom"):
        side = Side(style="thin")
        for c in range(1, cols+1):
            cl = ws.cell(r, c)
            b  = cl.border
            kw = {}
            if "bottom" in sides: kw["bottom"] = side
            if "top"    in sides: kw["top"]    = side
            if "all"    in sides: kw = {"top":side,"bottom":side,"left":side,"right":side}
            cl.border = Border(**kw)

    COLS = 12   # SL, Product, QTY, Desc, Model, Brand, Spec, HSN, Price/Pc, Amount, GST%, GST Val
    col_widths = [5, 20, 6, 22, 16, 12, 28, 11, 13, 14, 7, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ref_no      = quotation.get("ref_no", "—")
    client_name = quotation.get("client_name", "—")
    date_str    = datetime.now().strftime("%d %b %Y")
    valid_str   = (datetime.now().replace(day=datetime.now().day)).strftime("%d %b %Y")

    # ── Row 1: Company name ──
    ws.merge_cells(f"A1:L1")
    cell(1, 1, "DEMO", bold=True, size=18, color=WHITE, bg=NAVY, align="center")
    ws.row_dimensions[1].height = 30

    # ── Row 2: Sub-title ──
    ws.merge_cells("A2:L2")
    cell(2, 1, "Demo Company  |  123, Industrial Area, New Delhi – 110001  |  +91-98765-43210  |  info@demo.com",
         size=9, color=WHITE, bg=NAVY, align="center", italic=True)
    ws.row_dimensions[2].height = 16

    # ── Row 3: spacer ──
    ws.row_dimensions[3].height = 6

    # ── Row 4: QUOTATION title ──
    ws.merge_cells("A4:L4")
    cell(4, 1, "Q U O T A T I O N", bold=True, size=14, color=WHITE, bg=DARK, align="center")
    ws.row_dimensions[4].height = 24

    # ── Row 5: spacer ──
    ws.row_dimensions[5].height = 6

    # ── Rows 6-8: Meta info ──
    cell(6, 1, "Bill & Ship To:", bold=True, size=9, color=NAVY)
    cell(7, 1, client_name, bold=True, size=11, color=NAVY)
    cell(8, 1, "Attn: Purchase Manager", size=9, color="666666", italic=True)

    cell(6,  8, "Quotation No.", bold=True, size=9, color=NAVY)
    cell(6,  9, ref_no, bold=True, size=9)
    cell(7,  8, "Date",          bold=True, size=9, color=NAVY)
    cell(7,  9, date_str, size=9)
    cell(8,  8, "Valid Until",   bold=True, size=9, color=NAVY)
    cell(8,  9, "30 days from date above", size=9)
    cell(6, 10, "", bg=LIGHT); cell(7, 10, "", bg=LIGHT); cell(8, 10, "", bg=LIGHT)

    for r in [6,7,8]:
        ws.row_dimensions[r].height = 16

    # ── Row 9: spacer ──
    ws.row_dimensions[9].height = 6

    # ── Row 10: Salutation ──
    ws.merge_cells("A10:L10")
    cell(10, 1, "Dear Sir/Ma'am,  We are pleased to submit our best quotation for your kind consideration.",
         size=9, color="333333", italic=True)
    ws.row_dimensions[10].height = 16

    # ── Row 11: spacer ──
    ws.row_dimensions[11].height = 4

    # ── Row 12: Table header ──
    headers = ["SL", "PRODUCT", "QTY", "DESCRIPTION", "MODEL NO", "BRAND",
               "SPECIFICATION", "HSN", "PRICE/PC", "AMOUNT", "GST%", "GST VALUE"]
    for i, h in enumerate(headers, 1):
        cell(12, i, h, bold=True, size=9, color=WHITE, bg=NAVY, align="center")
    border_row(12, COLS, "all")
    ws.row_dimensions[12].height = 18

    # ── Data rows ──
    sub_total = 0.0
    gst_total = 0.0
    row_num   = 13

    for idx, item in enumerate(items):
        price   = float(item.get("price_per_pc") or item.get("price") or 0)
        qty     = int(item.get("qty") or 0)
        gst_pct = float(item.get("gst_pct") or 18)
        cur     = item.get("price_currency", "INR")
        # If USD, note in price column
        price_label = price  # store numeric
        amount  = qty * price
        gst_val = amount * gst_pct / 100
        sub_total += amount
        gst_total += gst_val

        bg_row = "F2F2F2" if idx % 2 == 1 else WHITE

        cell(row_num, 1,  idx+1,                       bg=bg_row, align="center", size=9)
        cell(row_num, 2,  item.get("product",""),       bg=bg_row, bold=True, size=9, wrap=True)
        cell(row_num, 3,  qty,                          bg=bg_row, align="center", size=9)
        cell(row_num, 4,  item.get("description",""),   bg=bg_row, size=8, wrap=True)
        cell(row_num, 5,  item.get("model_no",""),      bg=bg_row, size=8, wrap=True)
        cell(row_num, 6,  item.get("brand",""),         bg=bg_row, size=9)
        cell(row_num, 7,  item.get("specification",""), bg=bg_row, size=8, wrap=True)
        cell(row_num, 8,  item.get("hsn_code",""),      bg=bg_row, size=8, align="center")
        pc = ws.cell(row_num, 9,  price_label)
        pc.number_format = f'{"$" if cur=="USD" else "₹"}#,##0.00'
        pc.font = Font(size=9, name="Calibri"); pc.fill = PatternFill("solid", start_color=bg_row)
        pc.alignment = Alignment(horizontal="right")
        am = ws.cell(row_num, 10, amount)
        am.number_format = f'{"$" if cur=="USD" else "₹"}#,##0.00'
        am.font = Font(size=9, name="Calibri"); am.fill = PatternFill("solid", start_color=bg_row)
        am.alignment = Alignment(horizontal="right")
        cell(row_num, 11, gst_pct/100, bg=bg_row, align="center", size=9, num_fmt="0%")
        gv = ws.cell(row_num, 12, gst_val)
        gv.number_format = "₹#,##0.00"
        gv.font = Font(size=9, name="Calibri"); gv.fill = PatternFill("solid", start_color=bg_row)
        gv.alignment = Alignment(horizontal="right")

        border_row(row_num, COLS, "bottom")
        ws.row_dimensions[row_num].height = 40
        row_num += 1

    # ── Spacer ──
    row_num += 1

    # ── Sub total row ──
    ws.merge_cells(f"A{row_num}:H{row_num}")
    cell(row_num, 1, "SUB TOTAL", bold=True, size=10, color=NAVY, bg=LIGHT, align="right")
    ws.cell(row_num, 9).value = ""; ws.cell(row_num, 9).fill = PatternFill("solid", start_color=LIGHT)
    am2 = ws.cell(row_num, 10, sub_total)
    am2.number_format = "₹#,##0.00"; am2.font = Font(bold=True, size=10, name="Calibri")
    am2.fill = PatternFill("solid", start_color=LIGHT); am2.alignment = Alignment(horizontal="right")
    ws.cell(row_num, 11).fill = PatternFill("solid", start_color=LIGHT)
    gv2 = ws.cell(row_num, 12, gst_total)
    gv2.number_format = "₹#,##0.00"; gv2.font = Font(bold=True, size=10, name="Calibri")
    gv2.fill = PatternFill("solid", start_color=LIGHT); gv2.alignment = Alignment(horizontal="right")
    ws.row_dimensions[row_num].height = 20
    row_num += 1

    # ── Grand total row ──
    grand = sub_total + gst_total
    ws.merge_cells(f"A{row_num}:I{row_num}")
    cell(row_num, 1, "GRAND TOTAL  (incl. GST)", bold=True, size=12, color=WHITE, bg=NAVY, align="right")
    gt = ws.cell(row_num, 10, grand)
    gt.number_format = "₹#,##0.00"
    gt.font = Font(bold=True, size=12, color=WHITE, name="Calibri")
    gt.fill = PatternFill("solid", start_color=NAVY); gt.alignment = Alignment(horizontal="right")
    ws.merge_cells(f"J{row_num}:L{row_num}")
    ws.cell(row_num, 11).fill = PatternFill("solid", start_color=NAVY)
    ws.cell(row_num, 12).fill = PatternFill("solid", start_color=NAVY)
    ws.row_dimensions[row_num].height = 24
    row_num += 2

    # ── Amount in words ──
    ws.merge_cells(f"A{row_num}:L{row_num}")
    # Simple words (just show numeric for now, can expand)
    cell(row_num, 1, f"Amount in Words:  ₹{grand:,.2f}  (inclusive of all taxes)",
         size=9, color="333333", italic=True)
    ws.row_dimensions[row_num].height = 16
    row_num += 2

    # ── Footer ──
    ws.merge_cells(f"A{row_num}:L{row_num}")
    cell(row_num, 1, "This is a computer-generated quotation.  Thank you for choosing Demo.",
         size=8, color=WHITE, bg=NAVY, align="center", italic=True)
    ws.row_dimensions[row_num].height = 16

    # Freeze panes below header
    ws.freeze_panes = "A13"

    out = EXPORTS_DIR / f"Quote_{quotation.get('ref_no','export').replace('/','-')}.xlsx"
    wb.save(str(out))
    return str(out)


# ── Helpers ───────────────────────────────────────────────────────────────────

def get_boq_context(conn, prompt: str = "", catalogs: list = None) -> str:
    # Fetch items — filter by selected catalogs if specified
    if catalogs:
        placeholders = ",".join("?" * len(catalogs))
        rows = conn.execute(
            f"SELECT * FROM boq_items WHERE file_name IN ({placeholders}) ORDER BY file_name, product",
            catalogs
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM boq_items ORDER BY file_name, product").fetchall()
    if not rows:
        return "No product catalog uploaded yet."

    # Extract keywords — letters only (no numbers), ≥3 chars, skip stop words
    stop_words = {"the", "and", "for", "we", "need", "want", "give", "get",
                  "our", "us", "me", "a", "an", "of", "in", "is", "are",
                  "also", "with", "some", "please", "nos", "nos.", "pcs"}
    keywords = [w.lower() for w in re.split(r'\W+', prompt)
                if len(w) >= 3 and w.isalpha() and w.lower() not in stop_words]

    def relevance(r):
        # Compare everything in lowercase for case-insensitive matching
        text = " ".join([
            (r['product'] or ''),
            (r['brand'] or ''),
            (r['description'] or ''),
            (r['specification'] or ''),
        ]).lower()
        return sum(1 for k in keywords if k in text)

    # Phase 1: find all directly matching rows (relevance > 0)
    matched_bases = set()
    for r in rows:
        if relevance(r) > 0:
            # Collect base name (first 2 words, uppercase) to pull in all variants
            base = ' '.join((r['product'] or '').upper().split()[:2])
            matched_bases.add(base)

    # Phase 2: include all rows that either matched OR share a base name with a match
    # This ensures ALL variants from ALL catalogs are included
    included = []
    for r in rows:
        base = ' '.join((r['product'] or '').upper().split()[:2])
        if relevance(r) > 0 or base in matched_bases:
            included.append(r)

    if not included:
        # No keyword match — return top 15 by recency as fallback
        included = conn.execute(
            "SELECT * FROM boq_items ORDER BY uploaded_at DESC LIMIT 15"
        ).fetchall()

    # Sort included: directly matched items first, then variants
    def sort_key(r):
        return (0 if relevance(r) > 0 else 1, (r['product'] or ''))
    included.sort(key=sort_key)

    # Deduplicate: if same product name exists with both ₹0 and non-zero price,
    # keep only the non-zero one. Case-insensitive comparison.
    seen_products = {}
    for r in included:
        key = (r['product'] or '').upper().strip()
        if key not in seen_products:
            seen_products[key] = r
        else:
            # Prefer non-zero price
            if (seen_products[key]['price'] or 0) == 0 and (r['price'] or 0) > 0:
                seen_products[key] = r
    included = list(seen_products.values())

    # Cap at 30 items to stay within Groq 12K TPM limit
    included = included[:30]

    # Build compact catalog lines — keep fields minimal to save tokens
    lines = []
    for r in included:
        spec = (r['specification'] or '')[:30].replace('\n', ' ')
        price = r['price'] if r['price'] else 0
        lines.append(
            f"{r['product']}|{r['brand'] or '-'}|{r['model_no'] or '-'}"
            f"|{spec}|HSN:{r['hsn_code'] or '-'}|INR:{price}|GST:{r['gst_pct'] or 18}%"
        )
    return "\n".join(lines)


def get_feedback_context(conn) -> str:
    rows = conn.execute(
        "SELECT f.rating, f.missing_items FROM feedback f "
        "ORDER BY f.created_at DESC LIMIT 30"
    ).fetchall()
    if not rows:
        return "No feedback yet."
    return "\n".join(f"- Rating: {r['rating']} | Issue: {r['missing_items']}" for r in rows)


def generate_ref_no() -> str:
    conn = get_db()
    count = conn.execute("SELECT COUNT(*) FROM quotations").fetchone()[0]
    conn.close()
    return f"SMI-{datetime.now().strftime('%Y%m')}-{count + 1:04d}"


def get_latest_template(conn):
    return conn.execute("SELECT * FROM templates ORDER BY uploaded_at DESC LIMIT 1").fetchone()


# ── API ───────────────────────────────────────────────────────────────────────

@app.get("/")
def index():
    return FileResponse(str(BASE / "static" / "index.html"))


@app.post("/api/scan-boq")
async def scan_boq(file: UploadFile = File(...)):
    """Preview file contents without saving to DB."""
    import traceback, tempfile
    if not file.filename.endswith((".xls", ".xlsx")):
        raise HTTPException(400, "Only .xls/.xlsx files allowed")
    try:
        suffix = ".xlsx" if file.filename.endswith(".xlsx") else ".xls"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            shutil.copyfileobj(file.file, tmp)
            tmp_path = tmp.name
        items, structure, img_map = parse_boq_excel(tmp_path, file.filename)
        os.unlink(tmp_path)
        return {
            "filename": file.filename,
            "columns_detected": list(structure.get("col_map", {}).keys()),
            "total_products": len(items),
            "images_found": len(img_map),
            "preview": items[:8],  # first 8 rows
        }
    except Exception as e:
        raise HTTPException(500, f"Scan error: {type(e).__name__}: {e}\n{traceback.format_exc()[-300:]}")


@app.post("/api/upload-boq")
async def upload_boq(file: UploadFile = File(...)):
    import traceback
    if not file.filename.endswith((".xls", ".xlsx")):
        raise HTTPException(400, "Only .xls/.xlsx files allowed")
    try:
        return await _upload_boq(file)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Upload error: {type(e).__name__}: {e}\n{traceback.format_exc()[-400:]}")

async def _upload_boq(file: UploadFile):

    dest = UPLOADS_DIR / file.filename
    with open(str(dest), "wb") as f:
        shutil.copyfileobj(file.file, f)

    items, structure, img_map = parse_boq_excel(str(dest), file.filename)

    conn = get_db()
    # Check if already uploaded
    existing = conn.execute("SELECT COUNT(*) FROM boq_items WHERE file_name=?", (file.filename,)).fetchone()[0]
    if existing:
        conn.close()
        return {"message": f"⚠️ '{file.filename}' is already in the catalog ({existing} products). Delete it first if you want to re-upload.", "already_exists": True}
    for item in items:
        conn.execute(
            "INSERT INTO boq_items (file_name,product,description,model_no,brand,"
            "specification,hsn_code,price,price_currency,gst_pct,image_path,sheet_name,uploaded_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (item["file_name"], item["product"], item["description"], item["model_no"],
             item["brand"], item["specification"], item["hsn_code"], item["price"],
             item.get("price_currency","INR"), item["gst_pct"], item["image_path"],
             item.get("sheet_name",""), item["uploaded_at"])
        )

    # Save template structure (supports xlsx only for template-based generation)
    if file.filename.endswith(".xlsx"):
        conn.execute(
            "INSERT INTO templates (file_name, file_path, structure_json, uploaded_at) VALUES (?,?,?,?)",
            (file.filename, str(dest), json.dumps(structure), datetime.now().isoformat())
        )

    conn.commit()
    total = conn.execute("SELECT COUNT(*) FROM boq_items").fetchone()[0]
    imgs_found = len(img_map)
    conn.close()

    return {
        "message": f"Uploaded '{file.filename}' — extracted {len(items)} products, "
                   f"{imgs_found} images found. Total catalog: {total} items.",
        "columns_detected": list(structure.get("col_map", {}).keys()),
        "images_found": imgs_found,
    }


@app.get("/api/boq-files")
def list_boq_files():
    conn = get_db()
    rows = conn.execute("SELECT file_name, COUNT(*) as count, MAX(uploaded_at) as uploaded_at FROM boq_items GROUP BY file_name ORDER BY uploaded_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.delete("/api/boq-files/{filename:path}")
def delete_boq_file(filename: str):
    conn = get_db()
    conn.execute("DELETE FROM boq_items WHERE file_name=?", (filename,))
    conn.commit()
    conn.close()
    # Also delete physical file if exists
    dest = UPLOADS_DIR / filename
    if dest.exists():
        dest.unlink()
    return {"message": f"'{filename}' removed from catalog."}

@app.get("/api/boq-items")
def list_boq_items():
    conn = get_db()
    rows = conn.execute("SELECT * FROM boq_items ORDER BY uploaded_at DESC").fetchall()
    conn.close()
    result = []
    for r in rows:
        item = dict(r)
        item["has_image"] = bool(item.get("image_path") and Path(item["image_path"]).exists())
        result.append(item)
    return result


@app.get("/api/product-image/{item_id}")
def product_image(item_id: int):
    conn = get_db()
    row = conn.execute("SELECT image_path FROM boq_items WHERE id=?", (item_id,)).fetchone()
    conn.close()
    if not row or not row["image_path"]:
        raise HTTPException(404, "No image")
    p = Path(row["image_path"])
    if not p.exists():
        raise HTTPException(404, "Image file not found")
    return Response(content=p.read_bytes(), media_type="image/png")


class VariantsRequest(BaseModel):
    prompt: str
    client_name: str = ""
    api_key: str = ""
    catalogs: list = []

class BuildQuotationRequest(BaseModel):
    client_name: str = ""
    items: list = []

class GenerateRequest(BaseModel):
    prompt: str
    client_name: str = ""
    api_key: str = ""
    catalogs: list = []  # list of file_name strings; empty = search all


@app.post("/api/generate")
def generate_quotation(req: GenerateRequest):
  try:
   return _generate(req)
  except HTTPException:
   raise
  except Exception as e:
   import traceback
   raise HTTPException(500, f"{type(e).__name__}: {e}\n{traceback.format_exc()[-600:]}")

GROQ_API_KEY_DEFAULT = os.environ.get("GROQ_API_KEY", "")

def _generate(req: GenerateRequest):
    api_key = req.api_key or os.environ.get("GROQ_API_KEY", "") or GROQ_API_KEY_DEFAULT
    if not api_key:
        raise HTTPException(400, "Groq API key not set. Please enter your key in the Generate tab.")

    conn = get_db()
    boq_ctx = get_boq_context(conn, req.prompt, catalogs=req.catalogs or [])
    feedback_ctx = get_feedback_context(conn)
    conn.close()

    client = Groq(api_key=api_key)

    system_prompt = f"""You are a quotation assistant. Return ONLY a valid json object, no explanation.

CATALOG (format: PRODUCT|BRAND|MODEL|SPEC|HSN|INR_PRICE|GST|SOURCE_FILE):
{boq_ctx}

CRITICAL RULES:
- ONLY include products that exist in the CATALOG above — NEVER invent products not listed
- If a requested product is NOT in the catalog, skip it entirely — do NOT include it with 0 price
- Search the catalog case-insensitively (e.g. "hand towel" matches "HAND TOWEL")
- Include ALL catalog variants that match (e.g. "HOT PLATE" and "HOT PLATE 304" are both valid)
- Use EXACT INR_PRICE, brand, model from catalog — never use 0 if catalog shows a real price

FEEDBACK:
{feedback_ctx}

Return json format:
{{"client_name":"...","items":[{{"sl_no":1,"product":"HOT PLATE","qty":5,"description":"","model_no":"","brand":"MELANGE","specification":"SS 202, LPG","hsn_code":"73211100","price_per_pc":66000,"gst_pct":18}}]}}"""

    try:
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Customer: {req.client_name}\nRequirements: {req.prompt}\nReturn json."}
            ],
            temperature=0.2,
            max_tokens=1200,
            response_format={"type": "json_object"},
        )
    except Exception as e:
        raise HTTPException(500, f"Groq API error: {str(e)}")

    raw = response.choices[0].message.content.strip()
    print("LLM RAW OUTPUT:", raw[:500])
    match = re.search(r"```(?:json)?\s*([\s\S]+?)```", raw)
    if match:
        raw = match.group(1).strip()

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise HTTPException(500, f"LLM returned invalid JSON: {e}\nRaw: {raw[:400]}")

    # Normalize: LLM sometimes uses "order", "products", "line_items" instead of "items"
    if "items" not in data:
        for alt in ["order", "products", "line_items", "quotation_items", "rows"]:
            if alt in data:
                data["items"] = data[alt]
                break
        else:
            data["items"] = []

    # Normalize item keys: "item"/"name" → "product", "quantity" → "qty"
    for item in data["items"]:
        if "product" not in item:
            for alt in ["item", "name", "product_name", "description"]:
                if alt in item:
                    item["product"] = item[alt]
                    break
        if "qty" not in item:
            for alt in ["quantity", "amount", "count", "nos"]:
                if alt in item:
                    item["qty"] = item[alt]
                    break

    # Override all item fields with actual catalog data — never trust LLM prices
    conn = get_db()
    final_items = []
    for item in data.get("items", []):
        product_name = item.get("product", "").upper().strip()

        # Exact match first, then partial — prefer non-zero price
        catalog_row = conn.execute(
            "SELECT * FROM boq_items WHERE UPPER(product) = ? "
            "ORDER BY CASE WHEN price > 0 THEN 0 ELSE 1 END, price DESC LIMIT 1",
            (product_name,)
        ).fetchone()
        if not catalog_row:
            catalog_row = conn.execute(
                "SELECT * FROM boq_items WHERE UPPER(product) LIKE ? "
                "ORDER BY CASE WHEN price > 0 THEN 0 ELSE 1 END, price DESC LIMIT 1",
                (f"{product_name}%",)
            ).fetchone()

        if not catalog_row:
            continue

        qty = int(item.get("qty") or 1)
        price = float(catalog_row["price"] or 0)
        gst_pct = float(catalog_row["gst_pct"] or 18)
        amount = qty * price
        gst_value = amount * gst_pct / 100

        final_items.append({
            "sl_no":         item.get("sl_no", len(final_items) + 1),
            "product":       catalog_row["product"],
            "qty":           qty,
            "description":   catalog_row["description"] or "",
            "model_no":      catalog_row["model_no"] or "",
            "brand":         catalog_row["brand"] or "",
            "specification": catalog_row["specification"] or "",
            "hsn_code":      catalog_row["hsn_code"] or "",
            "price_per_pc":  price,
            "price_currency": catalog_row["price_currency"] or "INR",
            "gst_pct":       gst_pct,
            "amount":        amount if (catalog_row["price_currency"] or "INR") == "INR" else 0,
            "gst_value":     gst_value if (catalog_row["price_currency"] or "INR") == "INR" else 0,
            "catalog_image_id": catalog_row["id"],
            "image_path":    catalog_row["image_path"] or "",
        })

    data["items"] = final_items

    ref_no = generate_ref_no()
    data["ref_no"] = ref_no
    if req.client_name and not data.get("client_name"):
        data["client_name"] = req.client_name
    data["date"] = datetime.now().strftime("%d-%m-%Y")

    cur = conn.execute(
        "INSERT INTO quotations (ref_no, client_name, items_json, status, created_at) VALUES (?,?,?,?,?)",
        (ref_no, data.get("client_name", ""), json.dumps(data), "draft", datetime.now().isoformat())
    )
    data["id"] = cur.lastrowid
    conn.commit()
    conn.close()
    return data


@app.post("/api/smart-generate")
def smart_generate(req: GenerateRequest):
    try:
        return _smart_generate(req)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        raise HTTPException(500, f"{type(e).__name__}: {e}\n{traceback.format_exc()[-600:]}")

def _smart_generate(req: GenerateRequest):
    api_key = req.api_key or os.environ.get("GROQ_API_KEY", "") or GROQ_API_KEY_DEFAULT
    if not api_key:
        raise HTTPException(400, "Groq API key required")

    # Step 1: LLM extracts product names + qty (tiny prompt, fast)
    try:
        groq_client = Groq(api_key=api_key)
        resp = groq_client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content":
                 "Extract product names and quantities from the customer requirement. "
                 "Return ONLY valid JSON: {\"items\":[{\"product\":\"iron\",\"qty\":30}]} "
                 "Keep product names simple and generic. Default qty to 1 if not stated. "
                 "Do not add any product not mentioned."},
                {"role": "user", "content": req.prompt}
            ],
            max_tokens=400, temperature=0.1
        )
        raw = resp.choices[0].message.content.strip()
        raw = re.sub(r"```[a-z]*\n?", "", raw).strip().rstrip("```")
        extracted = json.loads(raw).get("items", [])
    except Exception as e:
        raise HTTPException(500, f"Extraction error: {e}")

    # Step 2: For each extracted item, find all DB variants + score + pick best
    conn = get_db()
    result_items = []
    not_found    = []

    for item in extracted:
        kw  = item.get("product", "").upper().strip()
        qty = int(item.get("qty") or 1)
        if not kw:
            continue

        try:
            if req.catalogs:
                ph   = ",".join("?" * len(req.catalogs))
                rows = conn.execute(
                    f"SELECT * FROM boq_items WHERE UPPER(product) LIKE ? AND file_name IN ({ph})",
                    [f"%{kw}%"] + req.catalogs
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM boq_items WHERE UPPER(product) LIKE ?",
                    (f"%{kw}%",)
                ).fetchall()
            variants = [dict(r) for r in rows]
        except Exception:
            variants = []

        # Filter: keyword must match as a whole word, not just any substring
        # e.g. "table" should NOT match "mountable"
        def word_match(v):
            prod = (v.get("product") or "").upper()
            words = re.split(r'[\s\-/|,\.&]+', prod)
            return any(w == kw or w.startswith(kw) for w in words)

        word_matched = [v for v in variants if word_match(v)]
        variants = word_matched if word_matched else variants  # fallback to all if none

        # Prefer variants with price > 0; only keep zero-price if nothing else available
        priced = [v for v in variants if (v.get("price") or 0) > 0]
        variants = priced if priced else variants

        if not variants:
            not_found.append(item.get("product", ""))
            continue

        # Score: exact > compact match > has price > INR preferred
        def score_v(v):
            s  = 0
            pu = (v.get("product") or "").upper().strip()
            if pu == kw:
                s += 100                          # exact match
            elif kw in pu:
                # prefer compact names (kw is large fraction of name)
                ratio = len(kw) / max(len(pu), 1)
                s += 40 + int(ratio * 40)         # up to 80 for compact
            if (v.get("price") or 0) > 0:
                s += 60                           # has a real price
            if (v.get("price_currency") or "INR") == "INR":
                s += 5                            # slight INR preference as tiebreaker
            return s

        variants_sorted = sorted(variants, key=score_v, reverse=True)
        best = variants_sorted[0]

        result_items.append({
            "sl_no":        len(result_items) + 1,
            "product":      best.get("product", ""),
            "qty":          qty,
            "description":  best.get("description", ""),
            "model_no":     best.get("model_no", ""),
            "brand":        best.get("brand", ""),
            "specification":best.get("specification", ""),
            "hsn_code":     best.get("hsn_code", ""),
            "price_per_pc": float(best.get("price") or 0),
            "price_currency":best.get("price_currency", "INR"),
            "gst_pct":      float(best.get("gst_pct") or 18),
            "_variants":    variants_sorted,
            "_requested":   item.get("product", ""),
        })

    ref_no = f"QT-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    data   = {"ref_no": ref_no, "client_name": req.client_name,
              "items": result_items, "not_found": not_found}

    # Save to DB (strip internal _ keys)
    clean_items = [{k: v for k, v in i.items() if not k.startswith("_")} for i in result_items]
    data_db = {**data, "items": clean_items}
    cur = conn.execute(
        "INSERT INTO quotations (ref_no,client_name,items_json,status,created_at) VALUES (?,?,?,?,?)",
        (ref_no, req.client_name, json.dumps(data_db), "draft", datetime.now().isoformat())
    )
    data["id"] = cur.lastrowid
    conn.commit()
    conn.close()
    return data


@app.post("/api/variants")
def get_variants(req: VariantsRequest):
    """Step 1 of new flow: extract items from prompt, return ALL DB variants per item."""
    import traceback
    api_key = req.api_key or os.environ.get("GROQ_API_KEY", "") or GROQ_API_KEY_DEFAULT
    if not api_key:
        raise HTTPException(400, "Groq API key required")

    # LLM: extract product names + qty only (no catalog needed, very small prompt)
    try:
        client = Groq(api_key=api_key)
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content":
                 "Extract product names and quantities from the requirement. "
                 "Return ONLY valid JSON: {\"items\":[{\"product\":\"iron\",\"qty\":30}]} "
                 "Keep product names simple and generic. Default qty to 1 if not stated."},
                {"role": "user", "content": req.prompt}
            ],
            max_tokens=400, temperature=0.1
        )
        raw = response.choices[0].message.content.strip()
        raw = re.sub(r"```[a-z]*\n?", "", raw).strip().rstrip("```")
        extracted = json.loads(raw).get("items", [])
    except Exception as e:
        raise HTTPException(500, f"Extraction error: {e}\n{traceback.format_exc()[-300:]}")

    conn = get_db()
    groups = []
    for item in extracted:
        kw = item.get("product", "").upper().strip()
        qty = item.get("qty", 1)
        try:
            if req.catalogs:
                ph = ",".join("?" * len(req.catalogs))
                rows = conn.execute(
                    f"SELECT * FROM boq_items WHERE UPPER(product) LIKE ? AND file_name IN ({ph}) "
                    f"ORDER BY CASE WHEN price>0 THEN 0 ELSE 1 END, price DESC",
                    [f"%{kw}%"] + req.catalogs
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM boq_items WHERE UPPER(product) LIKE ? "
                    "ORDER BY CASE WHEN price>0 THEN 0 ELSE 1 END, price DESC",
                    (f"%{kw}%",)
                ).fetchall()
            variants = [dict(r) for r in rows]
        except Exception:
            variants = []

        groups.append({"requested": item.get("product",""), "qty": qty,
                        "variants": variants, "found": len(variants) > 0})
    conn.close()
    return {"groups": groups, "client_name": req.client_name}


@app.post("/api/build-quotation")
def build_quotation(req: BuildQuotationRequest):
    """Step 2: save user-selected variants as a proper quotation."""
    ref_no = f"QT-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    enriched = []
    for i, item in enumerate(req.items):
        enriched.append({
            "sl_no": i + 1,
            "product":       item.get("product", ""),
            "qty":           item.get("qty", 1),
            "description":   item.get("description", ""),
            "model_no":      item.get("model_no", ""),
            "brand":         item.get("brand", ""),
            "specification": item.get("specification", ""),
            "hsn_code":      item.get("hsn_code", ""),
            "price_per_pc":  float(item.get("price", 0) or item.get("price_per_pc", 0)),
            "price_currency":item.get("price_currency", "INR"),
            "gst_pct":       float(item.get("gst_pct", 18)),
        })
    data = {"ref_no": ref_no, "client_name": req.client_name, "items": enriched}
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO quotations (ref_no, client_name, items_json, status, created_at) VALUES (?,?,?,?,?)",
        (ref_no, req.client_name, json.dumps(data), "draft", datetime.now().isoformat())
    )
    data["id"] = cur.lastrowid
    conn.commit()
    conn.close()
    return data


@app.get("/api/quotations")
def list_quotations(status: str = None):
    conn = get_db()
    if status:
        rows = conn.execute("SELECT * FROM quotations WHERE status=? ORDER BY created_at DESC", (status,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM quotations ORDER BY created_at DESC").fetchall()
    conn.close()
    result = []
    for r in rows:
        q = dict(r)
        q["items_json"] = json.loads(q["items_json"])
        result.append(q)
    return result


class UpdateItemsRequest(BaseModel):
    items: list
    client_name: str = ""


@app.put("/api/quotations/{qid}")
def update_quotation(qid: int, req: UpdateItemsRequest):
    conn = get_db()
    row = conn.execute("SELECT * FROM quotations WHERE id=?", (qid,)).fetchone()
    if not row:
        raise HTTPException(404, "Not found")
    data = json.loads(row["items_json"])
    data["items"] = req.items
    if req.client_name:
        data["client_name"] = req.client_name
    for item in data["items"]:
        qty = int(item.get("qty") or 0)
        price = float(item.get("price_per_pc") or 0)
        gst_pct = float(item.get("gst_pct") or 18)
        item["amount"] = qty * price
        item["gst_value"] = item["amount"] * gst_pct / 100
    conn.execute("UPDATE quotations SET items_json=?, client_name=? WHERE id=?",
                 (json.dumps(data), data["client_name"], qid))
    conn.commit()
    conn.close()
    data["id"] = qid
    return data


@app.get("/api/download/{qid}")
def download_quotation(qid: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM quotations WHERE id=?", (qid,)).fetchone()
    tmpl = get_latest_template(conn)
    conn.close()
    if not row:
        raise HTTPException(404, "Not found")

    data = json.loads(row["items_json"])
    items = data.get("items", [])

    path = build_xls_minimal(data, items)

    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"Quote_{data.get('ref_no', qid)}.xlsx"
    )


class FeedbackRequest(BaseModel):
    quotation_id: int
    rating: str
    missing_items: str = ""


@app.post("/api/feedback")
def submit_feedback(req: FeedbackRequest):
    conn = get_db()
    conn.execute(
        "INSERT INTO feedback (quotation_id, rating, missing_items, created_at) VALUES (?,?,?,?)",
        (req.quotation_id, req.rating, req.missing_items, datetime.now().isoformat())
    )
    if req.rating == "good":
        conn.execute("UPDATE quotations SET status='approved' WHERE id=?", (req.quotation_id,))
    conn.commit()
    conn.close()
    return {"message": "Feedback saved. Thank you!" if req.rating == "good" else "Feedback recorded — we'll improve!"}


@app.post("/api/approve/{qid}")
def approve_quotation(qid: int):
    conn = get_db()
    conn.execute("UPDATE quotations SET status='approved' WHERE id=?", (qid,))
    conn.commit()
    conn.close()
    return {"message": "Approved and stored in repository."}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
