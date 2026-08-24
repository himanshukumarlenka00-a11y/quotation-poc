import re, os, tempfile
from datetime import datetime
import pandas as pd
import openpyxl
from app.images import _save_image_to_disk, _xlsx_sheet_images, _assign_images_by_span

_TAUGHT = {}   # header→boq-field map, refreshed per parse from column_mappings
from app.xls_converter import convert_xls_to_xlsx

try:
    from app.xls_image_extractor import extract_images_with_fallback as _xls_extract_images_fb
except Exception:
    _xls_extract_images_fb = None


# ── Dynamic Template Analyzer ─────────────────────────────────────────────────

def analyze_template(wb: openpyxl.Workbook) -> dict:
    """
    Dynamically detect the structure of ANY quotation Excel template.
    Returns structure dict describing where each section lives.
    """
    ws = wb.active
    structure = {
        "header_rows": [],       # rows above the items table (company info)
        "bill_to_row": None,     # row index of "BILL & SHIP TO"
        "client_row": None,      # row where client name goes
        "date_row": None,        # row where date is
        "date_col": None,        # column where date is
        "ref_row": None,         # row where REF NO is
        "salutation_rows": [],   # Dear Sir + intro text rows
        "table_header_row": None, # row with SL.NO / PRODUCT / QTY etc.
        "data_start_row": None,
        "data_end_row": None,    # last actual data row (estimated from template)
        "footer_rows": [],       # rows after data (totals, T&C, bank, signature)
        "col_map": {},           # normalized column name -> 1-based col index
        "max_col": 1,
    }

    nrows = ws.max_row
    ncols = ws.max_column

    def cell_val(r, c):
        v = ws.cell(r, c).value
        return str(v).strip() if v is not None else ""

    def row_text(r):
        return " ".join(cell_val(r, c) for c in range(1, ncols + 1)).strip()

    # Find table header row (contains SL.NO or equivalent)
    for r in range(1, min(nrows + 1, 40)):
        txt = row_text(r).upper()
        if any(k in txt for k in ["SL.NO", "S.NO", "SL NO", "SLNO", "SR.NO"]):
            structure["table_header_row"] = r
            # Map columns
            for c in range(1, ncols + 1):
                v = cell_val(r, c).upper()
                if v:
                    structure["col_map"][v] = c
            structure["data_start_row"] = r + 1
            structure["max_col"] = ncols
            break

    if not structure["table_header_row"]:
        return structure

    thr = structure["table_header_row"]

    # Scan above header row for sections
    for r in range(1, thr):
        txt = row_text(r).upper()
        if "BILL" in txt and "SHIP" in txt:
            structure["bill_to_row"] = r
            # Date is usually on the same row but far right
            for c in range(ncols, 0, -1):
                v = cell_val(r, c)
                if "DATE" in v.upper():
                    structure["date_row"] = r
                    structure["date_col"] = c
                    break
            structure["client_row"] = r + 1
        if "REF" in txt and ("NO" in txt or ":" in txt):
            structure["ref_row"] = r
        structure["header_rows"].append(r)

    # Find data end row and footer
    in_data = False
    for r in range(thr + 1, nrows + 1):
        txt = row_text(r).upper()
        first = cell_val(r, 1).upper()
        # Stop at totals/footer markers
        if any(marker in first for marker in ["TOTAL", "GST VALUE", "GRAND TOTAL", "TERMS", "BANK", "CHEQUE"]):
            structure["data_end_row"] = r - 1
            structure["footer_rows"] = list(range(r, nrows + 1))
            break
        in_data = True

    if in_data and not structure["data_end_row"]:
        structure["data_end_row"] = nrows

    return structure


# ── BOQ Parser ────────────────────────────────────────────────────────────────

def _taught_boq_cols():
    """Learned header→field mappings (the same Teach table the master import
    uses), translated to this parser's column keys — so teaching a heading
    once fixes BOQ reads too. Best-effort: an empty dict on any failure."""
    xlate = {"product": "product", "original_model": "model_no", "brand": "brand",
             "specification": "specification", "hsn_code": "hsn_code",
             "gst_pct": "gst_pct", "price_inr": "price_inr",
             "price_usd": "price_usd", "unit": "unit", "qty": "qty"}
    try:
        from app.db import get_db
        conn = get_db()
        rows = conn.execute("SELECT header_norm, field FROM column_mappings").fetchall()
        conn.close()
        return {r["header_norm"]: xlate[r["field"]] for r in rows if r["field"] in xlate}
    except Exception:
        return {}


def parse_boq_excel(filepath: str, filename: str):
    """
    Fully dynamic parser — detects structure from the file itself.
    Returns (items, structure_dict)
    """
    global _TAUGHT
    _TAUGHT = _taught_boq_cols()
    items = []
    structure = {}

    # ── For .xls files, try converting to .xlsx via real Excel first ──────────
    # Some .xls "quotation" exports have their embedded-picture position data
    # truncated in a way our own byte-level parser can't fully recover (see
    # xls_image_extractor.py's docstring) — but real Excel resolves/repairs
    # this correctly on open. Converting through Excel and reading the result
    # with the already-reliable .xlsx path gets every image, not a guess.
    # Falls back to the raw .xls parser below if Excel/COM isn't available.
    converted_temp_path = None
    if filepath.endswith(".xls"):
        fd, converted_temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        if convert_xls_to_xlsx(filepath, converted_temp_path):
            filepath = converted_temp_path
        else:
            try:
                os.unlink(converted_temp_path)
            except OSError:
                pass
            converted_temp_path = None

    # ── Image extraction → {sheet_index: [(anchor_row, anchor_col, image_hash)]} ──
    # Col-aware so images can be assigned to products by ROW-RANGE (span),
    # preferring the OUR IMAGE column — this fixes the old drift/mis-mapping.
    sheet_images = {}
    leftover_blips_by_sheet = {}
    if filepath.endswith(".xls") and _xls_extract_images_fb:
        try:
            exact, leftover_blips_by_sheet = _xls_extract_images_fb(filepath)
            for im in exact:
                h = _save_image_to_disk(im["data"])
                if h:
                    sheet_images.setdefault(im["sheet_index"], []).append((im["row"], im["col"], h))
        except Exception as e:
            print(f"xls image extraction failed (non-fatal): {e}")

    # openpyxl only works with .xlsx — use it for images/structure when possible
    # (this also now covers .xls files successfully converted above)
    if filepath.endswith(".xlsx"):
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            try:
                structure = analyze_template(wb)
            finally:
                wb.close()  # release the file handle — Windows can't delete an open file
        except Exception:
            structure = {}
        try:
            sheet_images = _xlsx_sheet_images(filepath)
        except Exception:
            sheet_images = {}

    def _cleanup_converted_temp():
        if converted_temp_path:
            try:
                os.unlink(converted_temp_path)
            except OSError:
                pass  # Windows may still hold a lock briefly — non-fatal

    # Load all sheets with pandas
    try:
        xl = pd.ExcelFile(filepath)
        sheet_names = xl.sheet_names
        xl.close()
    except Exception:
        _cleanup_converted_temp()
        return items, structure

    # Note: "OPTION" is handled specially (treated as an alternative variant of the
    # product above it), so it is NOT in skip_keywords.
    skip_keywords = {"TOTAL", "GRAND TOTAL", "GST VALUE", "ADD GST", "NAN", "",
                     "SUMMARY", "TERMS", "BANK", "SI NO", "SHEET NAME"}

    for sheet_pos, sheet_name in enumerate(sheet_names):
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        except Exception:
            continue

        # Detect header row for this sheet independently
        sheet_thr = None
        sheet_col_map = {}
        for i, row in df.iterrows():
            vals = [str(v).upper().strip() for v in row.values if pd.notna(v)]
            joined = " || ".join(vals)
            # A header row names a product/description column AND at least one
            # data field (price/qty/model/amount). Substring match so variants
            # like "PRODUCT NAME", "MATERIAL DESCRIPTION", "SL. NO." all work.
            # "ITEM" must match as a whole cell too — client BOQs write the
            # product column as just "ITEM" (e.g. "SI NO | ITEM | QTY").
            has_name  = any(s in joined for s in
                            ["PRODUCT", "ITEM NAME", "ITEM CODE", "MATERIAL DESCRI",
                             "DESCRIPTION"]) or "ITEM" in vals
            has_field = any(s in joined for s in
                            ["PRICE", "QTY", "QUANTITY", "MODEL", "AMOUNT", "RATE", "MRP"])
            # A header an admin has TAUGHT as "product" also marks the row —
            # that's what makes BOQ column names fixable without code changes.
            if not has_name and any(_TAUGHT.get(v) == "product" for v in vals):
                has_name = True
            if has_name and has_field:
                sheet_thr = i
                for ci_, v in enumerate(row.values):
                    if pd.notna(v) and str(v).strip():
                        sheet_col_map[str(v).upper().strip()] = ci_
                break

        if sheet_thr is None:
            continue

        # Use structure from openpyxl for first sheet if available
        if not structure:
            structure["table_header_row"] = sheet_thr + 1
            structure["col_map"] = {k: v+1 for k, v in sheet_col_map.items()}

        # Taught headers win over the built-in guesses — an admin's explicit
        # mapping must never lose to a substring heuristic.
        taught_cols = {}
        for hdr, idx in sheet_col_map.items():
            f = _TAUGHT.get(hdr)
            if f is not None and f not in taught_cols:
                taught_cols[f] = idx

        def find_col(*names, exclude=None):
            """Find column index by matching name substring. exclude= list of substrings to avoid."""
            for n in names:
                for key, idx in sheet_col_map.items():
                    if n in key:
                        if exclude and any(ex in key for ex in exclude):
                            continue
                        return idx
            return None

        def col(key, *names, exclude=None):
            t = taught_cols.get(key)
            return t if t is not None else find_col(*names, exclude=exclude)

        ci = {
            "product":       col("product", "PRODUCT NAME", "PRODUCT", "ITEM NAME",
                                      "MATERIAL DESCRIPTION", "MATERIAL DESCRI", "DESCRIPTION", "ITEM"),
            # "REQUIREMENT" catches client-specific naming like "OPM Requirement"
            # — a client's own project/company code prefixed onto "Requirement"
            # is a common pattern for a BOQ's quantity column, not just this one.
            "qty":           col("qty", "TOTAL QTY", "QTY", "QUANTITY", "REQUIREMENT"),
            "description":   find_col("DESCRIPTION", "DESC"),
            "model_no":      col("model_no", "MODEL NO", "MODEL", "ITEM CODE", "ITEM CODES"),
            "brand":         col("brand", "BRAND", "MAKE"),
            "specification": col("specification", "OUR SPECIFICATION", "SPECIFICATION", "SPEC"),
            "hsn_code":      col("hsn_code", "HSN"),
            # INR price: explicitly exclude any column containing USD.
            # "AMOUNT" is a last-resort fallback — some client templates use
            # it to mean a per-unit price rather than qty×price, and only
            # checking it after every dedicated PRICE/RATE name comes up
            # empty means it never overrides a real price column when one
            # exists (e.g. a file with both PRICE/PC and AMOUNT columns,
            # where AMOUNT genuinely is qty×price, still uses PRICE/PC).
            "price_inr":     col("price_inr", "PRICE/PC INR", "PRICE INR", "PRICE/PC", "PRICE", "RATE", "UNIT PRICE", "AMOUNT", exclude=["USD", "DOLLAR"]),
            # USD price: must contain USD or DOLLAR
            "price_usd":     col("price_usd", "PRICE/PC USD", "UNIT PRICE USD", "PRICE USD", "AMOUNT USD", "TOTAL USD"),
            "gst_pct":       col("gst_pct", "GST%", "GST"),
            # qty×price column — distinct from the unit price unless the
            # sheet only has AMOUNT (then they're the same cell). Needed by
            # the format-preserving revised-quotation writer.
            "amount":        find_col("AMOUNT", exclude=["USD", "DOLLAR"]),
        }

        if ci["product"] is None:
            continue

        # Detect the image columns and assign each product its image by row-span
        # (preferring the OUR IMAGE column over the reference image).
        our_col = ref_col = None
        for key, idx in sheet_col_map.items():
            if "OUR IMAGE" in key:
                our_col = idx
            elif key == "IMAGE" or "REF IMAGE" in key or "REFERENCE IMAGE" in key:
                ref_col = idx
        # NOTE: the sequential best-effort fallback (fallback_blips=...) was
        # tried and reverted — in practice it assigned clearly wrong photos
        # (e.g. a toothbrush for "Bearing Extractor") often enough that it's
        # not safe to show on a real quote. A missing image is always safer
        # than a wrong one, so only confirmed (anchored) matches are used.
        span_img, guessed_img_rows = _assign_images_by_span(
            df, sheet_thr, ci["product"], sheet_images.get(sheet_pos, []), our_col, ref_col)

        last_product_name = ""   # tracks the most recent real product for OPTION rows

        for df_row_idx, row in df.iloc[sheet_thr + 1:].iterrows():
            vals = row.values

            product = vals[ci["product"]] if ci["product"] is not None else None
            blank_name = (product is None or (isinstance(product, float) and pd.isna(product))
                          or not str(product).strip())

            if blank_name:
                # A blank product name may be a VARIANT row — an alternative
                # model / price / spec for the product on the row(s) above
                # (e.g. a second clock model, an alternate kettle). Keep it,
                # inheriting that product's name, only when it carries its own
                # data; otherwise it's just a spacer row and is skipped.
                def _has(key):
                    idx = ci.get(key)
                    if idx is None:
                        return False
                    v = vals[idx]
                    if v is None or (isinstance(v, float) and pd.isna(v)):
                        return False
                    return str(v).strip() not in ("", "0", "0.0", "nan")
                if last_product_name and (_has("model_no") or _has("specification")
                        or _has("price_inr") or _has("price_usd") or _has("price")):
                    product_str = last_product_name
                else:
                    continue
            else:
                product_str = str(product).strip()
                if product_str.upper() in skip_keywords:
                    continue
                # Section/category header rows (e.g. "WELDING TOOLS :-",
                # "FURNITURE :-") are labels, not products — skip them.
                if re.search(r':\s*-?\s*$', product_str):
                    continue
                # "OPTION" / "OPTION 1" rows = alternative variant of the product
                # above. Inherit that product's name so it becomes a variant.
                if product_str.upper().startswith("OPTION"):
                    if last_product_name:
                        product_str = last_product_name
                    else:
                        continue   # no preceding product to attach to
                else:
                    last_product_name = product_str
            # Skip rows that look like headers or section titles (no price)
            price_col = ci.get("price_inr") or ci.get("price_usd") or ci.get("price")
            if price_col is not None:
                price_val = vals[price_col]
                if isinstance(price_val, str) and not price_val.replace('.','').replace(',','').isdigit():
                    continue

            def g(key):
                idx = ci.get(key)
                if idx is None: return ""
                v = vals[idx]
                if isinstance(v, float) and pd.isna(v): return ""
                return str(v).strip()

            def gn(key, default=0.0):
                idx = ci.get(key)
                if idx is None: return default
                try:
                    v = vals[idx]
                    return float(v) if pd.notna(v) else default
                except: return default

            # Image assigned by row-span (see _assign_images_by_span) — robust to
            # tall merged rows and two image columns. image_path holds a content
            # hash; the actual bytes live on disk under data/images/. image_match
            # is "exact" (real anchor found in the file), "guess" (best-effort
            # fallback for files with truncated anchor data — see
            # extract_images_with_fallback), or "" (no image found at all).
            image_path = span_img.get(df_row_idx, "")
            image_match = "guess" if df_row_idx in guessed_img_rows else ("exact" if image_path else "")

            items.append({
                "file_name": filename,
                "product": product_str,
                "qty": gn("qty", 0) or 1,
                "description": g("description"),
                "model_no": g("model_no"),
                "brand": g("brand"),
                "specification": g("specification"),
                "hsn_code": g("hsn_code"),
                "price": gn("price_inr") or gn("price_usd") or gn("price"),
                "price_currency": "INR" if gn("price_inr") else ("USD" if gn("price_usd") else "INR"),
                "gst_pct": (lambda g: round(g * 100, 2) if g and g <= 1 else (g or 18.0))(gn("gst_pct")),
                "image_path": image_path,
                "image_match": image_match,
                "uploaded_at": datetime.now().isoformat(),
                "sheet_name": sheet_name,
                # Provenance for the revised-quotation writer: exactly which
                # cell in the SOURCE workbook each value came from, so new
                # prices can be written back without touching the format.
                "_src": {"sheet": sheet_name, "row": int(df_row_idx) + 1,
                         "price_col": ci.get("price_inr"),
                         "qty_col": ci.get("qty"),
                         "amount_col": ci.get("amount")},
            })

    _cleanup_converted_temp()
    return items, structure
