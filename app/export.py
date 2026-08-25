from copy import copy
from datetime import datetime
import os
import re
import tempfile
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from app.config import EXPORTS_DIR
from app.images import _image_file_path

COMPANY_TEMPLATE_PATH = Path(__file__).parent / "assets" / "quotation_template.xlsx"


# ── Template-Based XLS Builder ────────────────────────────────────────────────

def copy_cell_style(src, dst):
    """Copy formatting from one cell to another."""
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def build_xls_from_template(template_path: str, structure: dict,
                             quotation: dict, items: list) -> str:
    """
    Build output XLS by cloning the actual template file and filling in data.
    Preserves ALL formatting, merged cells, fonts, colors from original.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    thr = structure.get("table_header_row", 1)
    data_start = thr + 1
    footer_rows = structure.get("footer_rows", [])
    footer_start = min(footer_rows) if footer_rows else ws.max_row + 1

    # ── Update variable header fields ────────────────────────────────────────
    # Client name
    client_row = structure.get("client_row")
    if client_row:
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(client_row, c).value or "")
            if v.strip() and "BILL" not in v.upper() and "SHIP" not in v.upper():
                ws.cell(client_row, c).value = quotation.get("client_name", "")
                break

    # Date
    date_row = structure.get("date_row")
    date_col = structure.get("date_col")
    if date_row and date_col:
        cell = ws.cell(date_row, date_col)
        existing = str(cell.value or "")
        prefix = existing.split(":")[0] + ": " if ":" in existing else "DATE : "
        cell.value = prefix + quotation.get("date", datetime.now().strftime("%d-%m-%Y"))

    # Ref No
    ref_row = structure.get("ref_row")
    if ref_row:
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(ref_row, c).value or "")
            if "REF" in v.upper():
                # Replace just the value part after ":"
                ws.cell(ref_row, c).value = f"REF NO: {quotation.get('ref_no', '')}"
                break

    # ── Remove existing data rows (between header and footer) ────────────────
    rows_to_delete = list(range(data_start, footer_start))
    if rows_to_delete:
        ws.delete_rows(data_start, len(rows_to_delete))

    # Get style from original first data row (before deletion) as template
    # We already deleted, so we'll build style manually using header row as reference
    col_map = structure.get("col_map", {})

    def find_col_idx(*names):
        for n in names:
            for key, idx in col_map.items():
                if n in key:
                    return idx
        return None

    ci = {
        "sl":     find_col_idx("SL.NO", "S.NO", "SL NO", "SLNO"),
        "prod":   find_col_idx("PRODUCT"),
        "qty":    find_col_idx("QTY", "QUANTITY"),
        "desc":   find_col_idx("DESCRIPTION"),
        "model":  find_col_idx("MODEL NO", "MODEL"),
        "brand":  find_col_idx("BRAND"),
        "image":  find_col_idx("IMAGE"),
        "spec":   find_col_idx("SPECIFICATION", "SPEC"),
        "hsn":    find_col_idx("HSN"),
        "price":  find_col_idx("PRICE/PC", "PRICE", "RATE"),
        "amount": find_col_idx("AMOUNT"),
        "gst":    find_col_idx("GST%", "GST"),
        "gstval": find_col_idx("GST VALUE", "GST VAL"),
    }

    max_col = structure.get("max_col", 8)

    # Reference style from header row
    ref_row_style = {c: ws.cell(thr, c) for c in range(1, max_col + 1)}

    # ── Insert new data rows ──────────────────────────────────────────────────
    for idx, item in enumerate(items):
        r = data_start + idx
        ws.insert_rows(r)

        price = float(item.get("price_per_pc") or item.get("price") or 0)
        qty = int(item.get("qty") or 0)
        gst_pct = float(item.get("gst_pct") or 18)
        amount = qty * price
        gst_val = amount * gst_pct / 100

        data = {
            ci["sl"]:     item.get("sl_no", idx + 1),
            ci["prod"]:   item.get("product", ""),
            ci["qty"]:    qty,
            ci["desc"]:   item.get("description", ""),
            ci["model"]:  item.get("model_no", ""),
            ci["brand"]:  item.get("brand", ""),
            ci["image"]:  "",
            ci["spec"]:   item.get("specification", "").replace("\\n", "\n"),
            ci["hsn"]:    item.get("hsn_code", ""),
            ci["price"]:  price,
            ci["amount"]: amount,
            ci["gst"]:    gst_pct / 100,
            ci["gstval"]: gst_val,
        }

        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            copy_cell_style(ref_row_style[c], cell)
            cell.value = data.get(c, "")

        # Number formats for numeric columns
        for col_key, fmt in [("price", "₹#,##0.00"), ("amount", "₹#,##0.00"),
                              ("gst", "0%"), ("gstval", "₹#,##0.00")]:
            col_idx = ci.get(col_key)
            if col_idx:
                ws.cell(r, col_idx).number_format = fmt

        # Embed product image if available
        img_file = _image_file_path(item.get("image_path", ""), full=True)
        if img_file and ci.get("image"):
            try:
                xl_img = XLImage(str(img_file))
                xl_img.width = 80
                xl_img.height = 60
                cell_addr = f"{get_column_letter(ci['image'])}{r}"
                ws.add_image(xl_img, cell_addr)
            except Exception:
                pass

        ws.row_dimensions[r].height = 65

    # ── Update totals in footer ───────────────────────────────────────────────
    total_amount = sum(int(i.get("qty") or 0) * float(i.get("price_per_pc") or i.get("price") or 0) for i in items)
    total_gst = sum(int(i.get("qty") or 0) * float(i.get("price_per_pc") or i.get("price") or 0) * float(i.get("gst_pct") or 18) / 100 for i in items)
    grand_total = total_amount + total_gst

    # Find and update total cells in footer (scan for TOTAL keyword)
    new_footer_start = data_start + len(items)
    for r in range(new_footer_start, ws.max_row + 1):
        for c in range(1, max_col + 1):
            v = str(ws.cell(r, c).value or "").upper().strip()
            if v == "TOTAL":
                # Put total amount in amount column
                if ci.get("amount"):
                    ws.cell(r, ci["amount"]).value = total_amount
                    ws.cell(r, ci["amount"]).number_format = "₹#,##0.00"
                if ci.get("gstval"):
                    ws.cell(r, ci["gstval"]).value = total_gst
                    ws.cell(r, ci["gstval"]).number_format = "₹#,##0.00"
            elif v in ("GST VALUE", "ADD GST@18%", "ADD GST"):
                if ci.get("amount"):
                    ws.cell(r, ci["amount"]).value = total_gst
                    ws.cell(r, ci["amount"]).number_format = "₹#,##0.00"
            elif "GRAND TOTAL" in v:
                if ci.get("amount"):
                    ws.cell(r, ci["amount"]).value = grand_total
                    ws.cell(r, ci["amount"]).number_format = "₹#,##0.00"

    out = EXPORTS_DIR / f"Quote_{quotation.get('ref_no','export').replace('/','-')}.xlsx"
    wb.save(str(out))
    return str(out)


def _embed_item_image(ws, img_file: Path, row: int, col: int, box_w: int = 250, box_h: int = 160,
                       center_height: int = None):
    """Embed a product photo into (row, col), scaled to fit inside box_w x box_h
    (pixels) while preserving aspect ratio. Centered horizontally in the cell
    and vertically within center_height (defaults to box_h) — pass the row's
    full pixel height there when box_h is capped smaller, so the image stays
    centered in tall rows instead of sitting near the top."""
    if center_height is None:
        center_height = box_h
    try:
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU
        from PIL import Image as PILImage
        import io as _io

        raw = img_file.read_bytes()
        pim = PILImage.open(_io.BytesIO(raw))
        ow, oh = pim.size
        scale = min(box_w / ow, box_h / oh)
        sw, sh = max(1, int(ow * scale)), max(1, int(oh * scale))

        xl_img = XLImage(_io.BytesIO(raw))
        marker = AnchorMarker(col=col - 1, colOff=pixels_to_EMU(max(2, (box_w - sw) // 2)),
                               row=row - 1, rowOff=pixels_to_EMU(max(2, (center_height - sh) // 2)))
        size = XDRPositiveSize2D(pixels_to_EMU(sw), pixels_to_EMU(sh))
        xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
        ws.add_image(xl_img)
    except Exception:
        pass


def _estimate_wrapped_lines(text: str, col_width_chars: float) -> int:
    """Rough estimate of how many wrapped lines `text` needs in a column of
    the given Excel character width — used to size row height so long
    specs/descriptions don't get visually clipped by a fixed row height."""
    if not text:
        return 1
    chars_per_line = max(8, int(col_width_chars - 1))
    total = 0
    for line in text.split("\n"):
        line = line.strip()
        total += max(1, -(-len(line) // chars_per_line)) if line else 1  # ceil div
    return max(1, total)


def build_company_quotation(quotation: dict, items: list) -> str:
    """Build the quotation Excel by cloning the real Shanti Metal Industries /
    Melange company template (app/assets/quotation_template.xlsx) — letterhead,
    logo, sales-contact block, terms & conditions and bank details all come
    from the template itself, untouched. Only the date, ref no, client name
    and the item rows (which vary per quote) are filled in.

    The footer block (freight line, totals, terms, bank/signatory) is
    snapshotted then rebuilt at its shifted row rather than moved in place
    with ws.insert_rows()/delete_rows() — those unreliably drop cell values
    on some columns when shifting a multi-row block, verified by testing."""
    wb = openpyxl.load_workbook(str(COMPANY_TEMPLATE_PATH))
    ws = wb["QUOTE"]

    ref_no = quotation.get("ref_no", "")
    client_name = quotation.get("client_name", "")
    date_str = datetime.now().strftime("%d.%m.%Y")

    # HSN codes are 8 digits — the template's I column is a shade too narrow
    # and wraps them mid-number.
    if (ws.column_dimensions["I"].width or 0) < 11:
        ws.column_dimensions["I"].width = 11

    ws["L9"] = f"DATE : {date_str}"
    ws["A17"] = f"REF NO: {ref_no}"
    # Bill & Ship To: the hand-written block from the quote (multi-line);
    # falls back to the bare client name for older quotes.
    bill_to = (quotation.get("bill_to") or "").strip()
    ws["A10"] = bill_to if bill_to else client_name
    if bill_to and "\n" in bill_to:
        from openpyxl.styles import Alignment
        ws["A10"].alignment = Alignment(wrap_text=True, vertical="top")

    # Sales concern block: the template ships it at J11-J13, which ends
    # mid-page — relocate to column M so it right-aligns with the DATE and
    # the table edge. Selected person overrides the template's default.
    from openpyxl.styles import Alignment
    sp = quotation.get("sales_person") or {}
    lines = ([f"SALES CONCERN PERSON : MR {sp['name']}",
              f"CONTACT N0: {sp.get('phone', '')}",
              f"MAIL ID: {sp.get('email', '')}"] if sp.get("name")
             else [ws["J11"].value, ws["J12"].value, ws["J13"].value])
    for row, text in zip((11, 12, 13), lines):
        ws[f"J{row}"] = None
        ws[f"M{row}"] = text
        ws[f"M{row}"].alignment = Alignment(horizontal="right")

    FIRST_ITEM_ROW = 21
    FOOTER_START = 22   # freight row onward in the template

    # BOQ Price / Profit are appended as new columns N/O rather than inserted
    # into the existing A..M layout — this template's footer already has
    # documented insert_rows()/delete_rows() corruption risk (see the
    # docstring above), and inserting COLUMNS through merged header/footer
    # ranges and the K/L/M formula references would carry the same risk.
    # Appending past the template's own columns touches none of that.
    has_boq_pricing = any(float(i.get("boq_price") or 0) > 0 for i in items)
    MAX_COL = 15 if has_boq_pricing else 13         # A..M, plus N..O when BOQ-priced
    BOQ_COL, PROFIT_COL = 14, 15                     # N, O

    if has_boq_pricing:
        # The letterhead block (rows 1-19: logo, sales-contact info) is boxed
        # with a medium border whose right edge sits on column M — appending
        # N/O past it would otherwise leave those two columns visually
        # outside that box while the table below them is wider.
        LETTERHEAD_LAST_ROW = 19

        # Rows that are a single full-row merge (e.g. A16:M16) render their
        # entire outline from the ANCHOR cell's (A's) border spec — every
        # other cell in the merge is a read-only MergedCell proxy that
        # mirrors the anchor for display and can't hold its own style.
        # Widening the merge itself (not editing any cell's border) is what
        # actually moves the right edge — verified directly: after widening
        # A16:M16 to A16:O16, O16 reports the anchor's medium right border
        # with no cell-level edit at all.
        row_merges = [rng for rng in list(ws.merged_cells.ranges)
                      if rng.min_row == rng.max_row and rng.min_row <= LETTERHEAD_LAST_ROW]
        merged_rows = {rng.min_row for rng in row_merges}
        for rng in row_merges:
            ws.unmerge_cells(str(rng))
            ws.merge_cells(start_row=rng.min_row, start_column=rng.min_col,
                            end_row=rng.max_row, end_column=PROFIT_COL)

        # Remaining (non-merged) rows genuinely store their own border per
        # cell — move the right edge from M to O and extend any top/bottom
        # divider lines (row 1's top, row 8/9's section break) across the
        # same width so the box reads as one continuous rectangle.
        for r in range(1, LETTERHEAD_LAST_ROW + 1):
            if r in merged_rows:
                continue
            row_ref = ws.cell(r, 12).border          # column L — this row's horizontal pattern
            m_border = ws.cell(r, 13).border
            top    = copy(row_ref.top) if row_ref.top and row_ref.top.style else None
            bottom = copy(row_ref.bottom) if row_ref.bottom and row_ref.bottom.style else None
            if m_border.right and m_border.right.style:
                ws.cell(r, 13).border = Border(left=m_border.left, top=m_border.top, bottom=m_border.bottom)
                ws.cell(r, PROFIT_COL).border = Border(right=copy(m_border.right), top=top, bottom=bottom)
            elif top or bottom:
                ws.cell(r, PROFIT_COL).border = Border(top=top, bottom=bottom)
            if top or bottom:
                ws.cell(r, BOQ_COL).border = Border(top=top, bottom=bottom)

        # The logo and sales-contact block were positioned for the box's
        # original (narrower) width — shift them right by the same amount
        # the box itself widened by, so they still sit in its corner instead
        # of leaving a stretch of empty space where the box grew.
        SHIFT = PROFIT_COL - 13   # = 2, matches the box's M->O widening
        for img in ws._images:
            anchor = getattr(img, "anchor", None)
            _from = getattr(anchor, "_from", None)
            if _from is not None and _from.col == 10 and _from.row == 1:   # the MELANGE logo
                _from.col += SHIFT
                # TwoCellAnchor images are positioned by BOTH corners — only
                # moving _from and leaving `to` in place collapses the box to
                # zero width, which is why the logo vanished on first try.
                to = getattr(anchor, "to", None)
                if to is not None:
                    to.col += SHIFT
        for coord in ("L9", "J11", "J12", "J13"):
            old_cell = ws[coord]
            new_cell = ws.cell(old_cell.row, old_cell.column + SHIFT)
            new_cell.value = old_cell.value
            copy_cell_style(old_cell, new_cell)
            old_cell.value = None

    # Drop the template's sample product photo (anchored at G21) — real
    # per-item photos are embedded fresh below. The logo (anchored near K2)
    # is left untouched.
    for img in list(ws._images):
        anchor = getattr(img, "anchor", None)
        _from = getattr(anchor, "_from", None)
        if _from is not None and _from.col == 6 and _from.row == FIRST_ITEM_ROW - 1:
            ws._images.remove(img)

    item_row_height = ws.row_dimensions[FIRST_ITEM_ROW].height
    spec_col_width = ws.column_dimensions["D"].width or 20
    desc_col_width = ws.column_dimensions["H"].width or 60
    LINE_HEIGHT_PT = 14.5
    ROW_PADDING_PT = 12
    MAX_ROW_HEIGHT_PT = 400  # cap for pathologically long specs (e.g. huge parts lists)

    # ── Snapshot the footer block (value + style + row height + merges)
    # before any mutation, then wipe it — it gets rebuilt below at its
    # shifted position. ──
    footer_last_row = ws.max_row
    footer_cells = {}
    for r in range(FOOTER_START, footer_last_row + 1):
        for c in range(1, MAX_COL + 1):
            cell = ws.cell(r, c)
            footer_cells[(r - FOOTER_START, c)] = {
                "value": cell.value,
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
            }
    footer_row_heights = {}
    for r in range(FOOTER_START, footer_last_row + 1):
        rd = ws.row_dimensions.get(r)
        footer_row_heights[r - FOOTER_START] = rd.height if rd else None
    footer_merges = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= FOOTER_START:
            footer_merges.append((rng.min_row - FOOTER_START, rng.min_col,
                                   rng.max_row - FOOTER_START, rng.max_col))
            ws.unmerge_cells(str(rng))

    ws.delete_rows(FOOTER_START, footer_last_row - FOOTER_START + 1)

    HEADER_ROW = 20
    if has_boq_pricing:
        # New header cells styled off the existing "AMOUNT" header (col M) —
        # matches the template's header look (navy fill, bold white text)
        # without hand-building a style from scratch.
        copy_cell_style(ws.cell(HEADER_ROW, 13), ws.cell(HEADER_ROW, BOQ_COL))
        copy_cell_style(ws.cell(HEADER_ROW, 13), ws.cell(HEADER_ROW, PROFIT_COL))
        ws.cell(HEADER_ROW, BOQ_COL, "BOQ PRICE")
        ws.cell(HEADER_ROW, PROFIT_COL, "PROFIT")
        ws.column_dimensions[get_column_letter(BOQ_COL)].width = 13
        ws.column_dimensions[get_column_letter(PROFIT_COL)].width = 13
        # Base style for the template's own first item row — idx>0 rows below
        # copy from this row, so it must be set before that loop runs.
        copy_cell_style(ws.cell(FIRST_ITEM_ROW, 13), ws.cell(FIRST_ITEM_ROW, BOQ_COL))
        copy_cell_style(ws.cell(FIRST_ITEM_ROW, 13), ws.cell(FIRST_ITEM_ROW, PROFIT_COL))

    # ── Item rows ──
    num_items = max(1, len(items))
    for idx, item in enumerate(items):
        r = FIRST_ITEM_ROW + idx
        if idx > 0:
            for c in range(1, MAX_COL + 1):
                copy_cell_style(ws.cell(FIRST_ITEM_ROW, c), ws.cell(r, c))

        price = float(item.get("price_per_pc") or item.get("price") or 0)
        qty = int(item.get("qty") or 0)
        # None-aware — a genuine 0% GST product must export as 0%, not
        # silently become 18% ("or 18" would treat 0 as missing).
        gst_pct = float(item.get("gst_pct")) if item.get("gst_pct") is not None else 18.0

        spec_text = (item.get("specification") or "").replace("\\n", "\n")
        # DESCRIPTION (col D) carries the client's own wording for the item —
        # the phrase they asked for — not a mirror of the master-table spec.
        desc_text = item.get("requested") or item.get("description") or ""

        # Row height must fit whichever of description/spec wraps to more
        # lines in its (much narrower) column — a fixed height clips long
        # text instead of growing to show it.
        text_lines = max(_estimate_wrapped_lines(desc_text, spec_col_width),
                          _estimate_wrapped_lines(spec_text, desc_col_width))
        row_height = max(item_row_height, min(text_lines * LINE_HEIGHT_PT + ROW_PADDING_PT, MAX_ROW_HEIGHT_PT))
        ws.row_dimensions[r].height = row_height

        ws.cell(r, 1, idx + 1)
        ws.cell(r, 2, item.get("product", ""))
        ws.cell(r, 3, qty)
        ws.cell(r, 4, desc_text)
        ws.cell(r, 5, item.get("model_no", ""))
        ws.cell(r, 6, item.get("brand", ""))
        ws.cell(r, 8, spec_text)
        ws.cell(r, 9, item.get("hsn_code", ""))
        # AMOUNT and GST AMOUNT are FORMULAS so the sheet stays alive: change
        # a qty, a price or a GST% in Excel and the line and every total below
        # follow. QTY / PRICE/PC / GST% are the only typed-in numbers.
        # (This used to be plain values, because a manual-calc install shows
        #  formula cells blank until F9 — wb.calculation.fullCalcOnLoad below
        #  forces the recalculation on open, which is what makes this safe.)
        amount = round(qty * price, 2)
        ws.cell(r, 10, round(price, 2))
        ws.cell(r, 11, f"=C{r}*J{r}")
        ws.cell(r, 12, gst_pct / 100)
        ws.cell(r, 13, f"=K{r}*L{r}")

        if has_boq_pricing:
            boq_price = float(item.get("boq_price") or 0)
            profit = qty * (boq_price - price)
            boq_cell = ws.cell(r, BOQ_COL, round(boq_price, 2))
            boq_cell.number_format = "₹#,##0.00"
            profit_cell = ws.cell(r, PROFIT_COL, round(profit, 2))
            profit_cell.number_format = "₹#,##0.00"
            profit_cell.font = Font(name=profit_cell.font.name, size=profit_cell.font.size,
                                     bold=profit_cell.font.bold,
                                     color="1E9E56" if profit >= 0 else "D64545")

        img_file = _image_file_path(item.get("image_path", ""), full=True)
        if img_file:
            # Cap the image's own size so it doesn't grow huge on long-text
            # rows, but center it within the row's real (uncapped) pixel
            # height so it doesn't sit near the top of a tall row.
            row_height_px = int(row_height * 96 / 72)
            img_box_h = min(int(item_row_height * 1.15), 220)
            _embed_item_image(ws, img_file, row=r, col=7,
                               box_w=250, box_h=img_box_h, center_height=row_height_px)

    # ── Rebuild the footer at its shifted position ──
    new_footer_start = FIRST_ITEM_ROW + num_items
    for (roff, c), snap in footer_cells.items():
        cell = ws.cell(new_footer_start + roff, c)
        cell.value = snap["value"]
        cell.font = snap["font"]
        cell.fill = snap["fill"]
        cell.border = snap["border"]
        cell.alignment = snap["alignment"]
        cell.number_format = snap["number_format"]
    for roff, height in footer_row_heights.items():
        if height is not None:
            ws.row_dimensions[new_footer_start + roff].height = height
    for roff_min, min_col, roff_max, max_col in footer_merges:
        ws.merge_cells(start_row=new_footer_start + roff_min, start_column=min_col,
                        end_row=new_footer_start + roff_max, end_column=max_col)

    freight_row = new_footer_start        # template offset 0 (orig row 22)
    total_row = new_footer_start + 1      # template offset 1 (orig row 23)
    gst_row = new_footer_start + 2        # template offset 2 (orig row 24)
    grand_row = new_footer_start + 3      # template offset 3 (orig row 25)

    freight_amt = float(quotation.get("freight_charge") or 0)
    freight_gst_pct = float(ws.cell(freight_row, 12).value or 0)   # template's GST% on the freight row (decimal)
    freight_gst = round(freight_amt * freight_gst_pct, 2)
    ws.cell(freight_row, 3, 1)
    ws.cell(freight_row, 10, freight_amt)
    ws.cell(freight_row, 11, f"=C{freight_row}*J{freight_row}")
    ws.cell(freight_row, 13, f"=K{freight_row}*L{freight_row}")

    total_amount = round(sum(
        round(int(i.get("qty") or 0) * float(i.get("price_per_pc") or i.get("price") or 0), 2)
        for i in items) + freight_amt, 2)
    total_gst = round(sum(
        round(round(int(i.get("qty") or 0) * float(i.get("price_per_pc") or i.get("price") or 0), 2)
              * (float(i.get("gst_pct")) if i.get("gst_pct") is not None else 18.0) / 100, 2)
        for i in items) + freight_gst, 2)
    # Totals sum the whole table, freight row included — its last row is
    # freight_row, which sits directly under the items.
    first, last = FIRST_ITEM_ROW, freight_row
    ws.cell(total_row, 11, f"=SUM(K{first}:K{last})")
    ws.cell(total_row, 13, f"=SUM(M{first}:M{last})")
    ws.cell(gst_row, 11, f"=SUM(M{first}:M{last})")
    ws.cell(grand_row, 11, f"=K{total_row}+K{gst_row}")

    if has_boq_pricing:
        # Same fix as the letterhead box: the totals block's right edge sits
        # on column M, unchanged by appending N/O — without moving it, the
        # profit total below reads as a disconnected floating box instead of
        # part of the same totals rectangle. Extend fill across N and move
        # the border (whichever weight each row uses) from M to O for every
        # row in this block, not just the one that gets a profit value.
        total_row_font = ws.cell(total_row, 13).font
        for r in (freight_row, total_row, gst_row, grand_row):
            m_cell = ws.cell(r, 13)
            m_border = m_cell.border
            ws.cell(r, BOQ_COL).fill = copy(m_cell.fill)
            if m_border.right and m_border.right.style:
                ws.cell(r, 13).border = Border(left=m_border.left, top=m_border.top, bottom=m_border.bottom)
                o_cell = ws.cell(r, PROFIT_COL)
                o_cell.fill = copy(m_cell.fill)
                o_cell.border = Border(right=copy(m_border.right), top=m_border.top, bottom=m_border.bottom)

        total_profit = sum(
            int(i.get("qty") or 0) * (float(i.get("boq_price") or 0) - float(i.get("price_per_pc") or i.get("price") or 0))
            for i in items
        )
        tp_cell = ws.cell(total_row, PROFIT_COL, round(total_profit, 2))
        tp_cell.number_format = "₹#,##0.00"
        tp_cell.font = Font(name=total_row_font.name, size=total_row_font.size, bold=True,
                             color="1E9E56" if total_profit >= 0 else "D64545")

    new_last_row = new_footer_start + (footer_last_row - FOOTER_START)
    ws.print_area = f"A1:{get_column_letter(MAX_COL)}{new_last_row}"

    # The template carries a literal SL "2" beside its PACKING, FORWARDING &
    # FREIGHT CHARGES row — a charges line is not an item, so its SL cell
    # stays blank on the final bill.
    for r in range(FIRST_ITEM_ROW, new_last_row + 1):
        for col in (2, 8):   # product (B) and specification (H) columns
            v = ws.cell(r, col).value
            if isinstance(v, str) and "PACKING, FORWARDING" in v.upper():
                ws.cell(r, 1).value = None
                break

    # openpyxl writes formulas with no cached result, so a workbook opened in
    # manual-calc mode would show every derived cell blank until someone hits
    # F9. This flag makes Excel recalculate the whole book on open, which is
    # what lets AMOUNT / GST / the totals be live formulas at all.
    wb.calculation.fullCalcOnLoad = True

    out = EXPORTS_DIR / f"Quote_{(ref_no or 'export').replace('/', '-')}.xlsx"
    wb.save(str(out))
    return str(out)


FINAL_BILL_TEMPLATE_PATH = Path(__file__).parent / "assets" / "final_bill_template.xlsx"

# Company bank block — constant on every outgoing bill. The company-format
# template carries it natively; the revised-copy export writes it under the
# client file's BANK DETAILS heading (their covers ship it empty).
BANK_DETAILS_LINES = [
    'CHEQUE PRINT NAME\xa0 \xa0 \xa0:\xa0SHANTI METAL INDUSTRIES',
    'PAYMENT MODE\xa0 \xa0 \xa0 \xa0 \xa0\xa0\xa0\xa0\xa0\xa0 :\xa0E-TRANSFER OR CHEQUE',
    'LOCATION\xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 :\xa0BANGALORE',
    'ACCOUNT NO\xa0 \xa0 \xa0\xa0\xa0\xa0\xa0\xa0\xa0\xa0 \xa0 \xa0 \xa0 \xa0: 2512011789',
    'IFSC CODE\xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0 \xa0: KKBK0000423',
    'BANK NAME\xa0 \xa0 \xa0 \xa0 \xa0\xa0\xa0\xa0\xa0\xa0\xa0\xa0\xa0\xa0 \xa0 \xa0 :\xa0KOTAK MAHINDRA BANK',
    'BANK ADDRESS\xa0 \xa0 \xa0 \xa0 \xa0\xa0\xa0\xa0\xa0\xa0\xa0\xa0 : BASAVANGUDI, BANGALORE - 560004',
]


def _norm_join(s):
    return re.sub(r"[^a-z0-9]+", "", str(s or "").lower())


def build_revised_from_source(quotation: dict, items: list, src_path: str) -> str:
    """Format-preserving revised quotation: the CLIENT'S OWN workbook is the
    template. Every sheet, header, image, box and column stays exactly as
    uploaded — only the numbers change: matched rows get the quote's price
    (and qty, if edited in the app), amounts and totals are recomputed as
    values, DATE becomes today, REF NO's version bumps (_V0 -> _V1) and the
    SUB line gains REVISED. Rows we could not match stay untouched."""
    from app.parser import parse_boq_excel
    src_rows, _ = parse_boq_excel(src_path, Path(src_path).name,
                                  skip_images=True)
    src_rows = [r for r in src_rows if r.get("_src")]
    if not src_rows:
        raise ValueError("no provenance rows in source")

    # .xls sources are written back through their .xlsx conversion
    work_path = src_path
    if src_path.lower().endswith(".xls"):
        from app.master_table import convert_xls_to_xlsx
        fd, conv = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        if not convert_xls_to_xlsx(src_path, conv):
            raise ValueError("xls source could not be converted")
        work_path = conv

    wb = openpyxl.load_workbook(work_path)
    wbv = openpyxl.load_workbook(work_path, data_only=True)

    # ── Align quote items to source rows (same order both came from; a
    # small lookahead + text check absorbs dropped/merged lines) ──
    written = {}
    wrote_money = False
    ptr = 0
    for row in src_rows:
        key = _norm_join(row.get("product")) + _norm_join(row.get("model_no"))
        hit = None
        for j in range(ptr, min(ptr + 6, len(items))):
            it = items[j]
            ik = _norm_join(it.get("requested") or it.get("_requested")
                            or it.get("product"))
            pk = _norm_join(row.get("product"))
            if pk and (pk in ik or ik in key or (ik and ik in pk)):
                hit = j
                break
        if hit is None:
            continue
        item = items[hit]
        ptr = hit + 1
        if item.get("not_in_catalog"):
            continue                      # unmatched: leave the row as-is
        srcp = row["_src"]
        sheet = srcp["sheet"]
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        r = srcp["row"]
        qty = int(item.get("qty") or 0) or int(row.get("qty") or 1)
        price = float(item.get("price_per_pc") or item.get("price") or 0)
        if price <= 0:
            continue
        pc, qc, ac = srcp.get("price_col"), srcp.get("qty_col"), srcp.get("amount_col")
        if pc is not None:
            ws.cell(r, pc + 1).value = round(price, 2)
            written[(sheet, r, pc + 1)] = round(price, 2)
            wrote_money = True
        if qc is not None:
            ws.cell(r, qc + 1).value = qty
            written[(sheet, r, qc + 1)] = qty
        if ac is not None and ac != pc:
            ws.cell(r, ac + 1).value = round(qty * price, 2)
            written[(sheet, r, ac + 1)] = round(qty * price, 2)
            wrote_money = True
    if not written or not wrote_money:
        # A price-less requirement BOQ (SL/PRODUCT/QTY only) has nowhere to
        # carry our prices — the company-format export is the right answer
        # there, so fail loudly and let the caller fall back.
        raise ValueError("source has no price/amount column to fill")

    # Manual freight/packing charge: written into the client's own
    # "ADD : PACKING , FREIGHT..." row BEFORE formulas recompute, so their
    # grand-total formula absorbs it. Untouched when not entered in the app.
    _fr = quotation.get("freight_charge")
    if _fr is not None:
        _fr = round(max(0.0, float(_fr)), 2)
        for sn in wb.sheetnames:
            ws_f = wb[sn]
            for row_cells in ws_f.iter_rows():
                for cell in row_cells:
                    v = cell.value
                    if (isinstance(v, str)
                            and re.search(r"PACKING|FREIGHT|FORWARDING", v, re.I)
                            and re.search(r"^\s*ADD\b|CHARGE", v, re.I)):
                        for cc in range(cell.column + 1, ws_f.max_column + 1):
                            t2 = ws_f.cell(cell.row, cc)
                            if t2.__class__.__name__ == "MergedCell":
                                continue
                            tv = t2.value
                            if (isinstance(tv, (int, float))
                                    or (isinstance(tv, str) and tv.startswith("="))):
                                t2.value = _fr
                                written[(sn, cell.row, cc)] = _fr
                                break
                        break

    # ── Recompute every formula we can into a VALUE (visible in Protected
    # View); anything unresolvable keeps its formula + calc-on-load ──
    from openpyxl.utils import range_boundaries, column_index_from_string

    def _val(sheet, r, c):
        v = written.get((sheet, r, c))
        if v is not None:
            return v
        cell = wb[sheet].cell(r, c).value
        if isinstance(cell, (int, float)):
            return cell
        if isinstance(cell, str) and not cell.startswith("="):
            return None
        cv = wbv[sheet].cell(r, c).value
        return cv if isinstance(cv, (int, float)) else None

    RESUM = re.compile(r"^=SUM\(\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)\)$", re.I)
    REREF = re.compile(r"'?([^'!=+\-*/(),]+?)'?!\$?([A-Z]{1,3})\$?(\d+)")
    unresolved = False
    for _pass in range(4):
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row_cells in ws.iter_rows():
                for cell in row_cells:
                    f = cell.value
                    if not (isinstance(f, str) and f.startswith("=")):
                        continue
                    m = RESUM.match(f.replace(" ", ""))
                    if m:
                        c1, r1, c2, r2 = (column_index_from_string(m.group(1)), int(m.group(2)),
                                          column_index_from_string(m.group(3)), int(m.group(4)))
                        vals = [_val(sn, rr, cc)
                                for rr in range(min(r1, r2), max(r1, r2) + 1)
                                for cc in range(min(c1, c2), max(c1, c2) + 1)]
                        cell.value = round(sum(v for v in vals if v is not None), 2)
                        written[(sn, cell.row, cell.column)] = cell.value
                        continue
                    # substitute cell references (cross-sheet first, then
                    # local), and evaluate what remains as plain arithmetic
                    expr = f[1:].replace(" ", "")
                    def _sub_x(mm):
                        v = _val(mm.group(1), int(mm.group(3)),
                                 column_index_from_string(mm.group(2)))
                        return "None" if v is None else repr(float(v))
                    expr = REREF.sub(_sub_x, expr)
                    def _sub_l(mm):
                        v = _val(sn, int(mm.group(2)),
                                 column_index_from_string(mm.group(1)))
                        return "None" if v is None else repr(float(v))
                    expr = re.sub(r"\$?([A-Z]{1,3})\$?(\d+)", _sub_l, expr)
                    if "None" not in expr and re.fullmatch(r"[\d.+\-*/() ]+", expr):
                        try:
                            cell.value = round(eval(expr, {"__builtins__": {}}), 2)
                            written[(sn, cell.row, cell.column)] = cell.value
                            continue
                        except Exception:
                            pass
                    unresolved = True
    if unresolved:
        wb.calculation.fullCalcOnLoad = True

    # ── Cover text: DATE -> today, REF version bump, SUB gains REVISED,
    # and the app's own BILL & SHIP TO + PREPARED BY reflect onto the
    # client's cover (replacing their TO block / filling under the date) ──
    today = datetime.now().strftime("%d-%m-%Y")
    _STOP = re.compile(r"^\s*(SUB\b|REF\b|DEAR\b|S\.?\s?NO|SL\.?\s?NO|TERMS|BANK|WE ARE)", re.I)
    bill_to = (quotation.get("bill_to") or "").strip() or (quotation.get("client_name") or "").strip()
    sp = quotation.get("sales_person") or {}
    for sn in wb.sheetnames:
        ws_c = wb[sn]
        date_cell = None
        for row_cells in ws_c.iter_rows():
            for cell in row_cells:
                v = cell.value
                if not isinstance(v, str):
                    continue
                if re.match(r"^\s*DATE\s*[:\-]", v, re.I):
                    cell.value = re.sub(r"(^\s*DATE\s*[:\-]\s*).*", rf"\g<1>{today}", v, flags=re.I)
                    date_cell = date_cell or cell
                elif re.search(r"\bREF\s*\.?\s*NO\b", v, re.I):
                    if re.search(r"_V(\d+)", v, re.I):
                        cell.value = re.sub(r"_V(\d+)",
                                            lambda m: f"_V{int(m.group(1)) + 1}", v)
                    else:
                        cell.value = v.rstrip() + "_V1"
                elif (re.match(r"^\s*SUB\b", v, re.I) and "QUOTATION" in v.upper()
                        and "REVISED" not in v.upper()):
                    cell.value = re.sub(r"(?i)QUOTATION", "REVISED QUOTATION", v, count=1)
                elif re.match(r"^\s*BANK\s*DETAILS?\s*[:\-]?\s*$", v, re.I):
                    # our bank block is constant on every bill, laid out
                    # like the reference: seven CONTIGUOUS lines in the
                    # label column, signatory off to the RIGHT. A client
                    # file with "Authorised Signatory" in the very next
                    # column split the block and clipped the overflow —
                    # such cells are shifted one column right first.
                    def _free(c):
                        return (c.__class__.__name__ != "MergedCell"
                                and (c.value is None
                                     or not str(c.value).strip()))
                    for k in range(1, 13):
                        nb = ws_c.cell(cell.row + k, cell.column + 1)
                        nv = nb.value
                        if (isinstance(nv, str)
                                and re.search(r"authorised\s+signat|^for\s+\w",
                                              nv.strip(), re.I)):
                            dst = ws_c.cell(cell.row + k, cell.column + 2)
                            if _free(dst):
                                dst.value = nv
                                copy_cell_style(nb, dst)
                                nb.value = None
                    rr = cell.row
                    for line in BANK_DETAILS_LINES:
                        placed = False
                        for _try in range(10):
                            rr += 1
                            tgt = ws_c.cell(rr, cell.column)
                            nxt = ws_c.cell(rr, cell.column + 1)
                            if _free(tgt) and _free(nxt):
                                tgt.value = line
                                if rr > cell.row + 1:
                                    copy_cell_style(
                                        ws_c.cell(cell.row + 1, cell.column), tgt)
                                placed = True
                                break
                        if not placed:
                            break
                elif (bill_to and re.match(r"^\s*(TO|BILL\s*&?\s*SHIP\s*TO)\s*[:\-]?\s*$",
                                            v, re.I)):
                    # write the app's client block into the rows below the
                    # TO label, stopping at the next section keyword
                    lines = [l.strip() for l in bill_to.splitlines() if l.strip()][:4]
                    r0 = cell.row
                    for k, line in enumerate(lines, 1):
                        tgt = ws_c.cell(r0 + k, cell.column)
                        tv = tgt.value
                        if tv is not None and (not isinstance(tv, str)
                                               or _STOP.match(tv)):
                            break
                        if tgt.__class__.__name__ == "MergedCell":
                            break
                        tgt.value = line
        # PREPARED BY under the date cell, only into genuinely empty cells
        if date_cell is not None and sp.get("name"):
            lines = [f"PREPARED BY : {sp['name']}"]
            if sp.get("phone"):
                lines.append(f"CONTACT NO : {sp['phone']}")
            for k, line in enumerate(lines, 1):
                tgt = ws_c.cell(date_cell.row + k, date_cell.column)
                if tgt.value is not None or tgt.__class__.__name__ == "MergedCell":
                    break
                tgt.value = line
                copy_cell_style(date_cell, tgt)

    ref_no = quotation.get("ref_no", "")
    out = EXPORTS_DIR / f"Quote_{(ref_no or 'revised').replace('/', '-')}.xlsx"
    wb.save(str(out))
    if work_path != src_path:
        try:
            os.unlink(work_path)
        except OSError:
            pass
    return str(out)


def build_final_bill(quotation: dict, items: list) -> str:
    """Build the final-bill workbook in the company's CYM-GWL design:
    a SUMMARY cover (letterhead + logo, BILL & SHIP TO, date, ref no,
    prepared-by, sheet-total index, T&C, bank details, signatory) plus a
    QUOTATION items sheet (S.No / Product / QTY / MODEL NO / BRAND / IMAGE /
    SPECIFICATION / PRICE/PC / AMOUNT) with embedded product photos and live
    formulas. The design comes from app/assets/final_bill_template.xlsx —
    cloned from the real client workbook — and is never restyled here."""
    wb = openpyxl.load_workbook(str(FINAL_BILL_TEMPLATE_PATH))
    ws_s, ws = wb["SUMMARY"], wb["QUOTATION"]

    ref_no = quotation.get("ref_no", "")
    date_str = datetime.now().strftime("%d-%m-%Y")

    # ── SUMMARY cover ──
    ws_s["C9"] = f"DATE : {date_str}"
    bill_to = (quotation.get("bill_to") or "").strip() or quotation.get("client_name", "")
    ws_s["A10"] = bill_to
    if "\n" in bill_to:
        ws_s["A10"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_s.row_dimensions[10].height = max(15, 13 * (bill_to.count("\n") + 1))
    ws_s["A13"] = f"REF NO: {ref_no}"
    sp = quotation.get("sales_person") or {}
    # C10/C11 are pre-styled underlined field boxes in the template.
    # Always fill PREPARED BY (team default matches the app's dropdown);
    # with no phone, open C11's box instead of rendering an empty one —
    # its top edge stays, doubling as PREPARED BY's underline.
    ws_s["C10"] = (f"PREPARED BY : MR {sp['name']}" if sp.get("name")
                   else "PREPARED BY : SALES TEAM")
    if sp.get("phone"):
        ws_s["C11"] = f"CONTACT NO : {sp['phone']}"
    else:
        b = ws_s["C11"].border
        ws_s["C11"].border = Border(right=b.right, top=b.top)

    # ── Items sheets — ONE PER SECTION when the quote came from a
    # multi-sheet upload, mirroring the source workbook's structure ──
    # r1 header, r2 item-row prototype, r3 totals-row prototype
    FIRST = 2
    totals_proto = 3
    totals_styles = [ws.cell(totals_proto, c)._style for c in range(1, 10)]
    spec_w = ws.column_dimensions["G"].width or 40

    def _fill_sheet(sheet, sheet_items):
        """Write one section's rows + totals into a template-styled sheet.
        Returns the sheet's amount total (computed VALUE — openpyxl can't
        store a formula's cached result, and formula cells render blank in
        Protected View, previews and unopened prints)."""
        sl = 0
        for idx, item in enumerate(sheet_items):
            r = FIRST + idx
            if idx > 0:
                for c in range(1, 10):
                    copy_cell_style(sheet.cell(FIRST, c), sheet.cell(r, c))
            product = item.get("product", "")
            is_charge = bool(re.search(r"packing|freight|forwarding",
                                        str(product), re.I))
            if not is_charge:
                sl += 1
                sheet.cell(r, 1, sl)
            qty = int(item.get("qty") or 0)
            price = float(item.get("price_per_pc") or item.get("price") or 0)
            spec_text = (item.get("specification") or "").replace("\\n", "\n")
            sheet.cell(r, 2, product)
            sheet.cell(r, 3, qty if qty else None)
            sheet.cell(r, 4, item.get("model_no", ""))
            sheet.cell(r, 5, item.get("brand", ""))
            sheet.cell(r, 7, spec_text)
            sheet.cell(r, 8, round(price, 2))
            sheet.cell(r, 9, round(qty * price, 2))
            # Compact rows — photo rows get room, text rows grow with spec
            img_file = _image_file_path(item.get("image_path", ""), full=True)
            lines = _estimate_wrapped_lines(spec_text, spec_w)
            base = 90 if img_file else 46
            sheet.row_dimensions[r].height = max(base, min(lines * 12 + 8, 190))
            if img_file:
                row_px = int(sheet.row_dimensions[r].height * 96 / 72)
                _embed_item_image(sheet, img_file, row=r, col=6,
                                   box_w=90, box_h=min(row_px - 6, 110),
                                   center_height=row_px)
        total_row = FIRST + len(sheet_items)
        for c in range(1, 10):
            sheet.cell(total_row, c)._style = totals_styles[c - 1]
        sheet.cell(total_row, 3, sum(int(i.get("qty") or 0) for i in sheet_items))
        tg = round(sum(int(i.get("qty") or 0)
                       * float(i.get("price_per_pc") or i.get("price") or 0)
                       for i in sheet_items), 2)
        sheet.cell(total_row, 9, tg)
        sheet.row_dimensions[total_row].height = 20
        sheet.print_area = f"A1:I{total_row}"
        return tg

    freight = max(0.0, float(quotation.get("freight_charge") or 0))

    sections = []
    for it in items:
        s = (it.get("section") or "").strip()
        if s not in sections:
            sections.append(s)
    multi = len([s for s in sections if s]) > 1

    if multi:
        # sanitized, unique Excel sheet names in source order
        used, names = set(), []
        for sec in sections:
            nm = re.sub(r"[\[\]:*?/\\]", " ", (sec or "ITEMS")).strip()[:31] or "ITEMS"
            base_nm, k = nm, 2
            while nm in used:
                nm = f"{base_nm[:27]}_{k}"
                k += 1
            used.add(nm)
            names.append(nm)
        sheets = [ws] + [wb.copy_worksheet(ws) for _ in names[1:]]
        for sheet, nm in zip(sheets, names):
            sheet.title = nm
        sheet_totals = [
            _fill_sheet(sheet, [i for i in items
                                if (i.get("section") or "").strip() == sec])
            for sheet, sec in zip(sheets, sections)]
        grand = round(sum(sheet_totals), 2)
        # SUMMARY index: one row per sheet (like the source workbook's own
        # cover), then TOTAL. Template ships with a single index row at 17.
        # openpyxl's insert_rows does NOT shift merged ranges — the TOTAL
        # row's A:B merge must be moved by hand or the new index cells are
        # read-only merge proxies that silently swallow writes.
        n = len(sheets)
        ws_s.unmerge_cells("A18:B18")
        ws_s.insert_rows(18, n - 1)
        for k in range(1, n):
            for c in range(1, 4):
                copy_cell_style(ws_s.cell(17, c), ws_s.cell(17 + k, c))
        for k, (nm, tg) in enumerate(zip(names, sheet_totals)):
            ws_s.cell(17 + k, 1, k + 1)
            ws_s.cell(17 + k, 2, nm)
            ws_s.cell(17 + k, 3, tg)
        ws_s.cell(17 + n, 3, grand)
        ws_s.merge_cells(start_row=17 + n, start_column=1,
                         end_row=17 + n, end_column=2)
        total_row_s = 17 + n
    else:
        grand = _fill_sheet(ws, items)
        ws_s["C17"] = grand
        ws_s["C18"] = grand
        total_row_s = 18
    if freight > 0:
        # Manual per-quote charge, exactly like the reference cover:
        # TOTAL / ADD : PACKING... / GRAND TOTAL. Two inserted rows styled
        # after the TOTAL row (inserting BELOW it leaves merges intact).
        ws_s.insert_rows(total_row_s + 1, 2)
        for k in (1, 2):
            for c in range(1, 4):
                copy_cell_style(ws_s.cell(total_row_s, c),
                                ws_s.cell(total_row_s + k, c))
        ws_s.merge_cells(start_row=total_row_s + 1, start_column=1,
                         end_row=total_row_s + 1, end_column=2)
        ws_s.merge_cells(start_row=total_row_s + 2, start_column=1,
                         end_row=total_row_s + 2, end_column=2)
        ws_s.cell(total_row_s + 1, 1,
                  "ADD : PACKING , FREIGHT AND FORWARDING CHARGES")
        ws_s.cell(total_row_s + 1, 3, round(freight, 2))
        ws_s.cell(total_row_s + 2, 1, "GRAND TOTAL")
        ws_s.cell(total_row_s + 2, 3, round(grand + freight, 2))
    out = EXPORTS_DIR / f"Quote_{(ref_no or 'export').replace('/', '-')}.xlsx"
    wb.save(str(out))
    return str(out)


USD_INR_RATE = 1.0  # USD column removed for now — prices shown as-is in INR (no conversion)

def get_usd_inr_rate() -> float:
    """USD column removed for now — no conversion (rate = 1)."""
    return USD_INR_RATE


def build_xls_minimal(quotation: dict, items: list) -> str:
    """Build a professional invoice-style Excel quotation (INR, with product images)."""
    import base64 as _b64, io as _io
    usd_rate = get_usd_inr_rate()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QUOTATION"

    # ── Colour palette ──
    NAVY   = "1F3864"
    GOLD   = "C8860C"
    LIGHT  = "DCE6F1"
    GREEN  = "E2EFDA"
    WHITE  = "FFFFFF"
    DARK   = "1F3864"

    def cell(r, c, val="", bold=False, size=10, color="000000", bg=None,
             align="left", valign="center", wrap=False, num_fmt=None, italic=False):
        cl = ws.cell(r, c, val)
        cl.font      = Font(bold=bold, size=size, color=color, name="Calibri", italic=italic)
        cl.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
        if bg:
            cl.fill = PatternFill("solid", start_color=bg)
        if num_fmt:
            cl.number_format = num_fmt
        return cl

    def border_row(r, cols, sides="bottom"):
        side = Side(style="thin")
        for c in range(1, cols+1):
            cl = ws.cell(r, c)
            b  = cl.border
            kw = {}
            if "bottom" in sides: kw["bottom"] = side
            if "top"    in sides: kw["top"]    = side
            if "all"    in sides: kw = {"top":side,"bottom":side,"left":side,"right":side}
            cl.border = Border(**kw)

    COLS = 12   # SL, Image, Product, QTY, Model, Brand, Spec, HSN, Price/Pc, Amount, GST%, GST Val
    col_widths = [5, 13, 20, 6, 15, 11, 26, 10, 13, 14, 7, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ref_no      = quotation.get("ref_no", "—")
    client_name = quotation.get("client_name", "—")
    date_str    = datetime.now().strftime("%d %b %Y")

    # ── Row 1: Company name ──
    ws.merge_cells("A1:L1")
    cell(1, 1, "DEMO", bold=True, size=18, color=WHITE, bg=NAVY, align="center")
    ws.row_dimensions[1].height = 30

    # ── Row 2: Sub-title ──
    ws.merge_cells("A2:L2")
    cell(2, 1, "Demo Company  |  123, Industrial Area, New Delhi – 110001  |  +91-98765-43210  |  info@demo.com",
         size=9, color=WHITE, bg=NAVY, align="center", italic=True)
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6

    # ── Row 4: QUOTATION title ──
    ws.merge_cells("A4:M4")
    cell(4, 1, "Q U O T A T I O N", bold=True, size=14, color=WHITE, bg=DARK, align="center")
    ws.row_dimensions[4].height = 24

    ws.row_dimensions[5].height = 6

    # ── Rows 6-8: Meta info ──
    cell(6, 1, "Bill & Ship To:", bold=True, size=9, color=NAVY)
    cell(7, 1, client_name, bold=True, size=11, color=NAVY)
    cell(8, 1, "Attn: Purchase Manager", size=9, color="666666", italic=True)

    cell(6,  9, "Quotation No.", bold=True, size=9, color=NAVY)
    cell(6, 10, ref_no, bold=True, size=9)
    cell(7,  9, "Date",          bold=True, size=9, color=NAVY)
    cell(7, 10, date_str, size=9)
    cell(8,  9, "Valid Until",   bold=True, size=9, color=NAVY)
    cell(8, 10, "30 days from date above", size=9)

    for r in [6,7,8]:
        ws.row_dimensions[r].height = 16

    ws.row_dimensions[9].height = 6

    # ── Row 10: Salutation ──
    ws.merge_cells("A10:M10")
    cell(10, 1, "Dear Sir/Ma'am,  We are pleased to submit our best quotation for your kind consideration.",
         size=9, color="333333", italic=True)
    ws.row_dimensions[10].height = 16

    ws.row_dimensions[11].height = 4

    # ── Row 12: Table header ──
    headers = ["SL", "IMAGE", "PRODUCT", "QTY", "DESCRIPTION", "MODEL NO", "BRAND",
               "SPECIFICATION", "HSN", "PRICE/PC (₹)", "AMOUNT (₹)", "GST%", "GST VALUE (₹)"]
    for i, h in enumerate(headers, 1):
        cell(12, i, h, bold=True, size=9, color=WHITE, bg=NAVY, align="center")
    border_row(12, COLS, "all")
    ws.row_dimensions[12].height = 18

    # ── Data rows ──
    sub_total = 0.0
    gst_total = 0.0
    row_num   = 13

    for idx, item in enumerate(items):
        price   = float(item.get("price_per_pc") or item.get("price") or 0)
        qty     = int(item.get("qty") or 0)
        gst_pct = float(item.get("gst_pct") or 18)
        cur     = item.get("price_currency", "INR")
        # Convert everything to INR
        if cur == "USD":
            price = price * usd_rate
        amount  = qty * price
        gst_val = amount * gst_pct / 100
        sub_total += amount
        gst_total += gst_val

        bg_row = "F2F2F2" if idx % 2 == 1 else WHITE

        cell(row_num, 1,  idx+1,                       bg=bg_row, align="center", size=9)
        cell(row_num, 2,  "",                           bg=bg_row)   # image placeholder
        cell(row_num, 3,  item.get("product",""),       bg=bg_row, bold=True, size=9, wrap=True)
        cell(row_num, 4,  qty,                          bg=bg_row, align="center", size=9)
        cell(row_num, 5,  item.get("description",""),   bg=bg_row, size=8, wrap=True)
        cell(row_num, 6,  item.get("model_no",""),      bg=bg_row, size=8, wrap=True)
        cell(row_num, 7,  item.get("brand",""),         bg=bg_row, size=9)
        cell(row_num, 8,  item.get("specification",""), bg=bg_row, size=8, wrap=True)
        cell(row_num, 9,  item.get("hsn_code",""),      bg=bg_row, size=8, align="center")
        pc = ws.cell(row_num, 10, round(price, 2))
        pc.number_format = '₹#,##0.00'
        pc.font = Font(size=9, name="Calibri"); pc.fill = PatternFill("solid", start_color=bg_row)
        pc.alignment = Alignment(horizontal="right")
        am = ws.cell(row_num, 11, round(amount, 2))
        am.number_format = '₹#,##0.00'
        am.font = Font(size=9, name="Calibri"); am.fill = PatternFill("solid", start_color=bg_row)
        am.alignment = Alignment(horizontal="right")
        cell(row_num, 12, gst_pct/100, bg=bg_row, align="center", size=9, num_fmt="0%")
        gv = ws.cell(row_num, 13, round(gst_val, 2))
        gv.number_format = "₹#,##0.00"
        gv.font = Font(size=9, name="Calibri"); gv.fill = PatternFill("solid", start_color=bg_row)
        gv.alignment = Alignment(horizontal="right")

        border_row(row_num, COLS, "bottom")
        ws.row_dimensions[row_num].height = 60   # ~80px tall box

        # ── Embed product image centered neatly inside the IMAGE cell ──
        img_file = _image_file_path(item.get("image_path", ""), full=True)
        if img_file:
            try:
                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                from openpyxl.drawing.xdr import XDRPositiveSize2D
                from openpyxl.utils.units import pixels_to_EMU
                from PIL import Image as PILImage

                raw = img_file.read_bytes()
                # Determine scaled size preserving aspect ratio inside box
                box_w, box_h = 76, 74          # px usable area inside the cell
                pim = PILImage.open(_io.BytesIO(raw))
                ow, oh = pim.size
                scale = min(box_w / ow, box_h / oh)
                sw, sh = max(1, int(ow * scale)), max(1, int(oh * scale))

                # center offsets within the cell (col B = index 1)
                col_px_w, row_px_h = 91, 80    # approx cell pixel size
                off_x = max(2, (col_px_w - sw) // 2)
                off_y = max(2, (row_px_h - sh) // 2)

                bio = _io.BytesIO(raw)
                xl_img = XLImage(bio)
                marker = AnchorMarker(col=1, colOff=pixels_to_EMU(off_x),
                                      row=row_num - 1, rowOff=pixels_to_EMU(off_y))
                size   = XDRPositiveSize2D(pixels_to_EMU(sw), pixels_to_EMU(sh))
                xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(xl_img)
            except Exception:
                pass

        row_num += 1

    row_num += 1

    # ── Sub total row ──
    ws.merge_cells(f"A{row_num}:I{row_num}")
    cell(row_num, 1, "SUB TOTAL", bold=True, size=10, color=NAVY, bg=LIGHT, align="right")
    am2 = ws.cell(row_num, 11, round(sub_total, 2))
    am2.number_format = "₹#,##0.00"; am2.font = Font(bold=True, size=10, name="Calibri")
    am2.fill = PatternFill("solid", start_color=LIGHT); am2.alignment = Alignment(horizontal="right")
    ws.cell(row_num, 10).fill = PatternFill("solid", start_color=LIGHT)
    ws.cell(row_num, 12).fill = PatternFill("solid", start_color=LIGHT)
    gv2 = ws.cell(row_num, 13, round(gst_total, 2))
    gv2.number_format = "₹#,##0.00"; gv2.font = Font(bold=True, size=10, name="Calibri")
    gv2.fill = PatternFill("solid", start_color=LIGHT); gv2.alignment = Alignment(horizontal="right")
    ws.row_dimensions[row_num].height = 20
    row_num += 1

    # ── Grand total row ──
    grand = sub_total + gst_total
    ws.merge_cells(f"A{row_num}:J{row_num}")
    cell(row_num, 1, "GRAND TOTAL  (incl. GST)", bold=True, size=12, color=WHITE, bg=NAVY, align="right")
    gt = ws.cell(row_num, 11, round(grand, 2))
    gt.number_format = "₹#,##0.00"
    gt.font = Font(bold=True, size=12, color=WHITE, name="Calibri")
    gt.fill = PatternFill("solid", start_color=NAVY); gt.alignment = Alignment(horizontal="right")
    ws.merge_cells(f"L{row_num}:M{row_num}")
    ws.cell(row_num, 12).fill = PatternFill("solid", start_color=NAVY)
    ws.cell(row_num, 13).fill = PatternFill("solid", start_color=NAVY)
    ws.row_dimensions[row_num].height = 24
    row_num += 2

    # ── Amount in words ──
    ws.merge_cells(f"A{row_num}:M{row_num}")
    cell(row_num, 1, f"Amount in Words:  {_amount_in_words(grand)}",
         size=9, color="333333", italic=True)
    ws.row_dimensions[row_num].height = 16
    row_num += 2

    # ── Footer ──
    ws.merge_cells(f"A{row_num}:M{row_num}")
    cell(row_num, 1, "This is a computer-generated quotation.  Thank you for choosing Demo.",
         size=8, color=WHITE, bg=NAVY, align="center", italic=True)
    ws.row_dimensions[row_num].height = 16

    ws.freeze_panes = "A13"

    out = EXPORTS_DIR / f"Quote_{quotation.get('ref_no','export').replace('/','-')}.xlsx"
    wb.save(str(out))
    return str(out)


def _amount_in_words(n: float) -> str:
    """Indian number system amount-in-words."""
    ones = ['','One','Two','Three','Four','Five','Six','Seven','Eight','Nine','Ten',
            'Eleven','Twelve','Thirteen','Fourteen','Fifteen','Sixteen','Seventeen',
            'Eighteen','Nineteen']
    tens = ['','','Twenty','Thirty','Forty','Fifty','Sixty','Seventy','Eighty','Ninety']
    def two(x):
        if x < 20: return ones[x]
        return tens[x//10] + ((' ' + ones[x%10]) if x%10 else '')
    def three(x):
        if x < 100: return two(x)
        return ones[x//100] + ' Hundred' + ((' ' + two(x%100)) if x%100 else '')
    n = int(round(n))
    if n == 0: return "Zero Rupees Only"
    parts = []
    crore = n // 10000000; n %= 10000000
    lakh  = n // 100000;   n %= 100000
    thou  = n // 1000;     n %= 1000
    hund  = n
    if crore: parts.append(three(crore) + ' Crore')
    if lakh:  parts.append(three(lakh) + ' Lakh')
    if thou:  parts.append(three(thou) + ' Thousand')
    if hund:  parts.append(three(hund))
    return ' '.join(parts) + ' Rupees Only'
